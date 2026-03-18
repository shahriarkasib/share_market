"""
DSE Historical Analysis Export
Generates an Excel file with 4 sheets:
  1. Sector Monthly Returns (sector seasonality)
  2. Stock Monthly Returns  (per-stock seasonality)
  3. MACD Floors            (per-stock MACD extremes + forward returns)
  4. MACD Extremes by Sector (sector-level summary)

Uses only A-category stocks from the fundamentals table.
Output: /Users/shariarsourav/Desktop/DSE_Historical_Analysis.xlsx
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from collections import defaultdict
from statistics import median
from database import get_connection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/shariarsourav/Desktop/DSE_Historical_Analysis.xlsx"
MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


def _style_header(ws, ncols):
    """Apply header styling to the first row."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def _auto_width(ws):
    """Set column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


# ── Data loading ────────────────────────────────────────────────────

def load_a_cat_symbols(conn):
    """Return dict {symbol: (sector, company_name)} for A-category stocks."""
    rows = conn.execute(
        "SELECT symbol, sector, company_name FROM fundamentals WHERE category = %s",
        ("A",),
    ).fetchall()
    return {r["symbol"]: (r["sector"] or "Unknown", r["company_name"] or r["symbol"]) for r in rows}


def load_daily_prices(conn, symbols):
    """Load daily prices for given symbols, ordered by symbol+date."""
    # Use ANY(array) for efficient IN-clause
    rows = conn.execute(
        "SELECT symbol, date, close, volume FROM daily_prices "
        "WHERE symbol = ANY(%s) AND close IS NOT NULL "
        "ORDER BY symbol, date",
        (list(symbols),),
    ).fetchall()
    # Group by symbol
    by_symbol = defaultdict(list)
    for r in rows:
        by_symbol[r["symbol"]].append(r)
    return by_symbol


def load_daily_analysis(conn, symbols):
    """Load daily_analysis rows (macd_hist) for given symbols."""
    rows = conn.execute(
        "SELECT symbol, date, macd_hist FROM daily_analysis "
        "WHERE symbol = ANY(%s) AND macd_hist IS NOT NULL "
        "ORDER BY symbol, date",
        (list(symbols),),
    ).fetchall()
    by_symbol = defaultdict(list)
    for r in rows:
        by_symbol[r["symbol"]].append(r)
    return by_symbol


# ── Sheet 1 & 2: Seasonality ───────────────────────────────────────

def compute_monthly_returns(prices_by_symbol, sym_info):
    """
    For each stock, compute monthly returns:
      return = (last close of month - first close of month) / first close of month * 100

    Returns:
      stock_monthly: list of (symbol, sector, month_num, return_pct, volume)
    """
    stock_monthly = []

    for symbol, rows in prices_by_symbol.items():
        if symbol not in sym_info:
            continue
        sector = sym_info[symbol][0]

        # Group by (year, month)
        by_ym = defaultdict(list)
        for r in rows:
            d = r["date"]
            by_ym[(d.year, d.month)].append(r)

        for (year, month), month_rows in by_ym.items():
            if len(month_rows) < 2:
                continue
            first_close = month_rows[0]["close"]
            last_close = month_rows[-1]["close"]
            if first_close is None or last_close is None or first_close == 0:
                continue
            ret = (last_close - first_close) / first_close * 100
            avg_vol = sum(r["volume"] or 0 for r in month_rows) / len(month_rows)
            stock_monthly.append((symbol, sector, month, ret, avg_vol, year))

    return stock_monthly


def build_sector_seasonality(stock_monthly):
    """Aggregate monthly returns by sector and month."""
    # key: (sector, month) -> list of (return_pct, avg_vol)
    agg = defaultdict(list)
    for symbol, sector, month, ret, vol, year in stock_monthly:
        agg[(sector, month)].append((ret, vol))

    rows = []
    for (sector, month) in sorted(agg.keys()):
        returns = [d[0] for d in agg[(sector, month)]]
        volumes = [d[1] for d in agg[(sector, month)]]
        n = len(returns)
        if n == 0:
            continue
        avg_ret = sum(returns) / n
        med_ret = median(returns)
        win_rate = sum(1 for r in returns if r > 0) / n * 100
        avg_vol = sum(volumes) / n
        rows.append({
            "sector": sector,
            "month": MONTH_NAMES[month - 1],
            "month_num": month,
            "avg_return_pct": round(avg_ret, 2),
            "median_return_pct": round(med_ret, 2),
            "positive_months_pct": round(win_rate, 1),
            "avg_volume": int(avg_vol),
            "sample_count": n,
        })

    rows.sort(key=lambda r: (r["sector"], r["month_num"]))
    return rows


def build_stock_seasonality(stock_monthly):
    """Per-stock monthly seasonality."""
    # key: (symbol, month) -> list of (return_pct, year)
    agg = defaultdict(list)
    sym_sector = {}
    for symbol, sector, month, ret, vol, year in stock_monthly:
        agg[(symbol, month)].append(ret)
        sym_sector[symbol] = sector

    rows = []
    for (symbol, month) in sorted(agg.keys()):
        data = agg[(symbol, month)]
        n = len(data)
        if n == 0:
            continue
        avg_ret = sum(data) / n
        med_ret = median(data)
        win_rate = sum(1 for r in data if r > 0) / n * 100
        rows.append({
            "symbol": symbol,
            "sector": sym_sector.get(symbol, "Unknown"),
            "month": MONTH_NAMES[month - 1],
            "month_num": month,
            "avg_return_pct": round(avg_ret, 2),
            "median_return_pct": round(med_ret, 2),
            "win_rate_pct": round(win_rate, 1),
            "sample_years": n,
        })

    rows.sort(key=lambda r: (r["symbol"], r["month_num"]))
    return rows


# ── Sheet 3 & 4: MACD Floor Analysis ──────────────────────────────

def build_macd_floors(analysis_by_symbol, prices_by_symbol, sym_info):
    """
    For each stock:
      - Find minimum macd_hist ever and its date
      - Forward returns 5d/10d/20d from that date
      - Current macd_hist (latest)
      - Percentage of current vs floor within the stock's range
      - Average bounce from extreme lows (bottom 10th percentile)
    """
    results = []

    for symbol in sorted(analysis_by_symbol.keys()):
        if symbol not in sym_info:
            continue
        sector = sym_info[symbol][0]
        analysis_rows = analysis_by_symbol[symbol]
        if len(analysis_rows) < 2:
            continue

        # All macd_hist values
        hist_values = [r["macd_hist"] for r in analysis_rows]
        min_hist = min(hist_values)
        max_hist = max(hist_values)
        current_hist = analysis_rows[-1]["macd_hist"]

        # Find the date of the minimum
        min_row = min(analysis_rows, key=lambda r: r["macd_hist"])
        min_date = min_row["date"]

        # Build price lookup by date for this symbol
        price_rows = prices_by_symbol.get(symbol, [])
        if not price_rows:
            continue
        price_dates = [r["date"] for r in price_rows]
        price_by_date = {r["date"]: r["close"] for r in price_rows}

        # Find price at min_date (or nearest trading day after)
        price_at_min = None
        idx_at_min = None
        for i, d in enumerate(price_dates):
            if d >= min_date:
                price_at_min = price_rows[i]["close"]
                idx_at_min = i
                break
        if price_at_min is None or price_at_min == 0:
            continue

        # Forward returns
        def _fwd_price(base_idx, days):
            target = base_idx + days
            if target < len(price_rows):
                return price_rows[target]["close"]
            return None

        p5 = _fwd_price(idx_at_min, 5)
        p10 = _fwd_price(idx_at_min, 10)
        p20 = _fwd_price(idx_at_min, 20)

        ret5 = round((p5 - price_at_min) / price_at_min * 100, 2) if p5 else None
        ret10 = round((p10 - price_at_min) / price_at_min * 100, 2) if p10 else None
        ret20 = round((p20 - price_at_min) / price_at_min * 100, 2) if p20 else None

        # current vs floor pct: 0% = at floor, 100% = at max
        hist_range = max_hist - min_hist
        if hist_range == 0:
            current_vs_floor_pct = 50.0
        else:
            current_vs_floor_pct = round((current_hist - min_hist) / hist_range * 100, 1)

        # Average bounce from extreme lows (bottom 10th percentile of macd_hist)
        sorted_hists = sorted(hist_values)
        p10_threshold = sorted_hists[max(0, int(len(sorted_hists) * 0.10) - 1)]
        extreme_dates = [r["date"] for r in analysis_rows if r["macd_hist"] <= p10_threshold]

        bounces = []
        for ed in extreme_dates:
            # Find price index for this date
            for i, d in enumerate(price_dates):
                if d >= ed:
                    base_price = price_rows[i]["close"]
                    fwd = _fwd_price(i, 10)
                    if base_price and fwd and base_price > 0:
                        bounces.append((fwd - base_price) / base_price * 100)
                    break

        avg_bounce = round(sum(bounces) / len(bounces), 2) if bounces else None

        results.append({
            "symbol": symbol,
            "sector": sector,
            "current_macd_hist": round(current_hist, 4),
            "min_macd_hist_ever": round(min_hist, 4),
            "min_macd_date": str(min_date),
            "price_at_min": round(price_at_min, 1),
            "price_5d_after": round(p5, 1) if p5 else None,
            "price_10d_after": round(p10, 1) if p10 else None,
            "price_20d_after": round(p20, 1) if p20 else None,
            "return_5d_pct": ret5,
            "return_10d_pct": ret10,
            "return_20d_pct": ret20,
            "current_vs_floor_pct": current_vs_floor_pct,
            "avg_bounce_from_extreme_lows_pct": avg_bounce,
        })

    results.sort(key=lambda r: r["current_vs_floor_pct"])
    return results


def build_macd_sector_summary(macd_floors):
    """Group MACD floor data by sector."""
    by_sector = defaultdict(list)
    for r in macd_floors:
        by_sector[r["sector"]].append(r)

    rows = []
    for sector in sorted(by_sector.keys()):
        stocks = by_sector[sector]
        n = len(stocks)
        avg_min = sum(s["min_macd_hist_ever"] for s in stocks) / n

        b5 = [s["return_5d_pct"] for s in stocks if s["return_5d_pct"] is not None]
        b10 = [s["return_10d_pct"] for s in stocks if s["return_10d_pct"] is not None]
        b20 = [s["return_20d_pct"] for s in stocks if s["return_20d_pct"] is not None]

        near_floor = sum(1 for s in stocks if s["current_vs_floor_pct"] <= 20)

        rows.append({
            "sector": sector,
            "stock_count": n,
            "avg_min_macd": round(avg_min, 4),
            "avg_bounce_5d": round(sum(b5) / len(b5), 2) if b5 else None,
            "avg_bounce_10d": round(sum(b10) / len(b10), 2) if b10 else None,
            "avg_bounce_20d": round(sum(b20) / len(b20), 2) if b20 else None,
            "stocks_near_floor_count": near_floor,
        })

    rows.sort(key=lambda r: r["sector"])
    return rows


# ── Excel writing ──────────────────────────────────────────────────

def write_sheet(ws, headers, rows, key_map):
    """Write headers + data rows to a worksheet."""
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, len(headers))

    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(key_map, 1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(key))

    _auto_width(ws)


def main():
    print("Connecting to database...")
    conn = get_connection()

    try:
        print("Loading A-category symbols...")
        sym_info = load_a_cat_symbols(conn)
        symbols = list(sym_info.keys())
        print(f"  Found {len(symbols)} A-category stocks")

        print("Loading daily prices...")
        prices_by_symbol = load_daily_prices(conn, symbols)
        print(f"  Loaded prices for {len(prices_by_symbol)} stocks")

        print("Loading daily analysis (MACD data)...")
        analysis_by_symbol = load_daily_analysis(conn, symbols)
        print(f"  Loaded analysis for {len(analysis_by_symbol)} stocks")

        # ── Seasonality ──
        print("Computing monthly returns...")
        stock_monthly = compute_monthly_returns(prices_by_symbol, sym_info)
        print(f"  {len(stock_monthly)} stock-month observations")

        print("Building sector seasonality...")
        sector_season = build_sector_seasonality(stock_monthly)

        # Add best-month summary rows per sector
        sectors_seen = set()
        best_months = {}
        for r in sector_season:
            s = r["sector"]
            if s not in best_months or r["avg_return_pct"] > best_months[s]["avg_return_pct"]:
                best_months[s] = r
        # Append summary rows at end
        summary_rows = []
        for s, best in sorted(best_months.items()):
            summary_rows.append({
                "sector": f"** {s} BEST **",
                "month": best["month"],
                "month_num": best["month_num"],
                "avg_return_pct": best["avg_return_pct"],
                "median_return_pct": best["median_return_pct"],
                "positive_months_pct": best["positive_months_pct"],
                "avg_volume": best["avg_volume"],
                "sample_count": best["sample_count"],
            })
        sector_season_full = sector_season + summary_rows

        print("Building stock seasonality...")
        stock_season = build_stock_seasonality(stock_monthly)

        # ── MACD Floors ──
        print("Building MACD floor analysis...")
        macd_floors = build_macd_floors(analysis_by_symbol, prices_by_symbol, sym_info)
        print(f"  {len(macd_floors)} stocks with MACD data")

        print("Building MACD sector summary...")
        macd_sector = build_macd_sector_summary(macd_floors)

        # ── Write Excel ──
        print("Writing Excel file...")
        wb = Workbook()

        # Sheet 1: Sector Monthly Returns
        ws1 = wb.active
        ws1.title = "Sector Monthly Returns"
        write_sheet(ws1,
            ["Sector", "Month", "Avg Return %", "Median Return %",
             "Win Rate %", "Avg Volume", "Sample Count"],
            sector_season_full,
            ["sector", "month", "avg_return_pct", "median_return_pct",
             "positive_months_pct", "avg_volume", "sample_count"],
        )

        # Sheet 2: Stock Monthly Returns
        ws2 = wb.create_sheet("Stock Monthly Returns")
        write_sheet(ws2,
            ["Symbol", "Sector", "Month", "Avg Return %",
             "Median Return %", "Win Rate %", "Sample Years"],
            stock_season,
            ["symbol", "sector", "month", "avg_return_pct",
             "median_return_pct", "win_rate_pct", "sample_years"],
        )

        # Sheet 3: MACD Floors
        ws3 = wb.create_sheet("MACD Floors")
        write_sheet(ws3,
            ["Symbol", "Sector", "Current MACD Hist", "Min MACD Hist Ever",
             "Min MACD Date", "Price at Min", "Price 5d After", "Price 10d After",
             "Price 20d After", "Return 5d %", "Return 10d %", "Return 20d %",
             "Current vs Floor %", "Avg Bounce from Extreme Lows %"],
            macd_floors,
            ["symbol", "sector", "current_macd_hist", "min_macd_hist_ever",
             "min_macd_date", "price_at_min", "price_5d_after", "price_10d_after",
             "price_20d_after", "return_5d_pct", "return_10d_pct", "return_20d_pct",
             "current_vs_floor_pct", "avg_bounce_from_extreme_lows_pct"],
        )

        # Sheet 4: MACD Extremes by Sector
        ws4 = wb.create_sheet("MACD Extremes by Sector")
        write_sheet(ws4,
            ["Sector", "Stock Count", "Avg Min MACD", "Avg Bounce 5d %",
             "Avg Bounce 10d %", "Avg Bounce 20d %", "Stocks Near Floor (<=20%)"],
            macd_sector,
            ["sector", "stock_count", "avg_min_macd", "avg_bounce_5d",
             "avg_bounce_10d", "avg_bounce_20d", "stocks_near_floor_count"],
        )

        wb.save(OUTPUT_PATH)
        print(f"\nDone! Saved to {OUTPUT_PATH}")
        print(f"  Sheet 1: {len(sector_season_full)} rows (sector seasonality + best-month summaries)")
        print(f"  Sheet 2: {len(stock_season)} rows (stock seasonality)")
        print(f"  Sheet 3: {len(macd_floors)} rows (MACD floors)")
        print(f"  Sheet 4: {len(macd_sector)} rows (MACD sector summary)")

    finally:
        conn.close()


if __name__ == "__main__":
    main()
