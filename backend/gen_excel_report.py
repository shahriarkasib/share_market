"""Generate comprehensive DSE portfolio analysis Excel report."""
import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

# ── DB connection ──
conn = psycopg2.connect(
    'postgresql://postgres.iihlezpkpllacztoaguc:160021062Ss%23%23@aws-1-ap-southeast-1.pooler.supabase.com:6543/postgres',
    cursor_factory=psycopg2.extras.RealDictCursor
)
cur = conn.cursor()

# ── Portfolio definition ──
portfolio = {
    'ORIONINFU': {'qty': 65, 'avg': 369.5},
    'ROBI':      {'qty': 850, 'avg': 30.7},
    'GP':        {'qty': 411, 'avg': 244.5},
    'HWAWELLTEX':{'qty': 1000, 'avg': 44.8},
}
watchlist = ['LOVELLO']
all_symbols = list(portfolio.keys()) + watchlist
DATE = '2026-03-10'

# ── Fetch data ──
daily = {}
llm = {}
judge = {}

for sym in all_symbols:
    cur.execute("SELECT * FROM daily_analysis WHERE date = %s AND symbol = %s", (DATE, sym))
    r = cur.fetchone()
    if r:
        daily[sym] = dict(r)

    cur.execute("SELECT * FROM llm_daily_analysis WHERE date = %s AND symbol = %s", (DATE, sym))
    r = cur.fetchone()
    if r:
        llm[sym] = {k: v for k, v in dict(r).items() if v is not None}

    cur.execute("SELECT * FROM judge_daily_analysis WHERE date = %s AND symbol = %s", (DATE, sym))
    r = cur.fetchone()
    if r:
        judge[sym] = {k: v for k, v in dict(r).items() if v is not None}

# DSEX
cur.execute("SELECT * FROM dsex_history ORDER BY date DESC LIMIT 1")
dsex = dict(cur.fetchone())

# Top buy signals
cur.execute("""SELECT symbol, action, score, entry_low, entry_high, sl, t1, t2, ltp, macd_status, rsi
              FROM daily_analysis WHERE date = '2026-03-10' AND action LIKE '%%BUY%%' AND action NOT LIKE '%%AVOID%%'
              ORDER BY score DESC LIMIT 20""")
top_buys = [dict(r) for r in cur.fetchall()]

# Top 5 buy LLM how_to_buy
top5_how = {}
for sig in top_buys[:5]:
    sym = sig['symbol']
    cur.execute("SELECT how_to_buy FROM llm_daily_analysis WHERE date = %s AND symbol = %s", (DATE, sym))
    r = cur.fetchone()
    if r and r.get('how_to_buy'):
        top5_how[sym] = r['how_to_buy']

conn.close()

# ── Helper to convert DB values ──
def fv(val):
    """Convert DB value to float, return None if not possible."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, Decimal):
        return float(val)
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def sv(val):
    """Convert to string safely."""
    if val is None:
        return ''
    return str(val)

# ── Styles ──
DARK_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
MED_BLUE = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
WHITE_FONT = Font(color='FFFFFF', bold=True, size=11)
WHITE_FONT_MED = Font(color='FFFFFF', bold=True, size=10)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
HEADER_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')

def style_header_row(ws, row, cols, fill=None, font=None):
    if fill is None:
        fill = DARK_BLUE
    if font is None:
        font = WHITE_FONT
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data_row(ws, row, cols, alt=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = NORMAL_FONT
        if alt:
            cell.fill = LIGHT_GRAY

def auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                longest = max(len(l) for l in lines)
                if longest > max_len:
                    max_len = longest
        w = min(max(max_len + 2, min_w), max_w)
        ws.column_dimensions[col_letter].width = w

# ── Workbook ──
wb = Workbook()

# ============================================================
# Sheet 1: PORTFOLIO DASHBOARD
# ============================================================
ws1 = wb.active
ws1.title = "PORTFOLIO DASHBOARD"

# Title
ws1.merge_cells('A1:K1')
ws1['A1'] = f'DSE Portfolio Dashboard — {DATE} (Tuesday)'
ws1['A1'].font = Font(bold=True, size=16, color='1F4E79')

ws1.merge_cells('A2:K2')
dsex_val = fv(dsex.get('close')) or fv(dsex.get('value'))
dsex_chg = fv(dsex.get('change'))
dsex_pct = fv(dsex.get('change_pct'))
dsex_str = f"DSEX: {dsex_val}"
if dsex_chg is not None:
    dsex_str += f" | Change: {dsex_chg:+.1f}"
if dsex_pct is not None:
    dsex_str += f" ({dsex_pct:+.2f}%)"
ws1['A2'] = dsex_str
ws1['A2'].font = Font(size=11, italic=True)

# Headers
headers = ['Symbol', 'Qty', 'Avg Cost', 'LTP', 'Change%', 'P/L (BDT)', 'Entry(AI)', 'SL(AI)', 'T1', 'T2', 'Action']
row = 4
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, len(headers))

# Data rows
total_invested = 0
total_current = 0
row = 5
for idx, (sym, info) in enumerate(portfolio.items()):
    d = daily.get(sym, {})
    ltp = fv(d.get('ltp'))
    qty = info['qty']
    avg = info['avg']
    invested = qty * avg
    current = qty * ltp if ltp else 0
    pl = current - invested
    total_invested += invested
    total_current += current

    change_pct_val = None
    if ltp and avg:
        change_pct_val = ((ltp - avg) / avg) * 100

    # Get action from judge if available, else daily
    action = ''
    j = judge.get(sym, {})
    if j.get('final_action'):
        action = sv(j['final_action'])
    elif d.get('action'):
        action = sv(d['action'])

    ws1.cell(row=row, column=1, value=sym)
    ws1.cell(row=row, column=2, value=qty)
    ws1.cell(row=row, column=3, value=round(avg, 1))
    ws1.cell(row=row, column=4, value=round(ltp, 1) if ltp else 'N/A')
    if change_pct_val is not None:
        ws1.cell(row=row, column=5, value=f"{change_pct_val:+.1f}%")
    else:
        ws1.cell(row=row, column=5, value='N/A')
    ws1.cell(row=row, column=6, value=round(pl, 0) if ltp else 'N/A')
    ws1.cell(row=row, column=7, value=f"{fv(d.get('entry_low')):.1f}-{fv(d.get('entry_high')):.1f}" if d.get('entry_low') else 'N/A')
    ws1.cell(row=row, column=8, value=round(fv(d.get('sl')), 1) if d.get('sl') else 'N/A')
    ws1.cell(row=row, column=9, value=round(fv(d.get('t1')), 1) if d.get('t1') else 'N/A')
    ws1.cell(row=row, column=10, value=round(fv(d.get('t2')), 1) if d.get('t2') else 'N/A')
    ws1.cell(row=row, column=11, value=action)

    alt = idx % 2 == 1
    style_data_row(ws1, row, len(headers), alt=alt)

    # Color P/L cell
    pl_cell = ws1.cell(row=row, column=6)
    if ltp:
        if pl > 0:
            pl_cell.fill = GREEN_FILL
        elif pl < 0:
            pl_cell.fill = RED_FILL

    row += 1

# Notes for specific stocks
notes_row = row + 1
ws1.merge_cells(f'A{notes_row}:K{notes_row}')
ws1.cell(row=notes_row, column=1, value='NOTES:')
ws1.cell(row=notes_row, column=1).font = BOLD_FONT

notes_row += 1
ws1.merge_cells(f'A{notes_row}:K{notes_row}')
ws1.cell(row=notes_row, column=1, value='* GP: Bought AFTER record date — NO dividend applies')
ws1.cell(row=notes_row, column=1).font = Font(size=10, color='CC0000')

notes_row += 1
ws1.merge_cells(f'A{notes_row}:K{notes_row}')
ws1.cell(row=notes_row, column=1, value='* HWAWELLTEX: T+2 locked until Thursday March 12, 2026')
ws1.cell(row=notes_row, column=1).font = Font(size=10, color='CC0000')

notes_row += 1
ws1.merge_cells(f'A{notes_row}:K{notes_row}')
ws1.cell(row=notes_row, column=1, value='* User does NOT buy bank or insurance stocks')
ws1.cell(row=notes_row, column=1).font = Font(size=10, color='666666')

# Summary section
sum_row = notes_row + 2
ws1.merge_cells(f'A{sum_row}:C{sum_row}')
ws1.cell(row=sum_row, column=1, value='PORTFOLIO SUMMARY')
ws1.cell(row=sum_row, column=1).font = Font(bold=True, size=13, color='1F4E79')

realized_pl = -1080  # MAGURAPLEX sold 600@80.5 vs 82.3 avg
unrealized_pl = total_current - total_invested
net_pl = unrealized_pl + realized_pl

summary_items = [
    ('Total Invested', f'{total_invested:,.0f} BDT'),
    ('Total Current Value', f'{total_current:,.0f} BDT'),
    ('Unrealized P/L', f'{unrealized_pl:+,.0f} BDT'),
    ('Realized P/L (MAGURAPLEX sold 600@80.5)', f'{realized_pl:+,.0f} BDT'),
    ('Net P/L', f'{net_pl:+,.0f} BDT'),
]

for i, (label, val) in enumerate(summary_items):
    r = sum_row + 1 + i
    ws1.cell(row=r, column=1, value=label)
    ws1.cell(row=r, column=1).font = BOLD_FONT
    ws1.cell(row=r, column=3, value=val)
    ws1.cell(row=r, column=3).font = NORMAL_FONT
    # Color net P/L
    if label == 'Net P/L':
        if net_pl > 0:
            ws1.cell(row=r, column=3).fill = GREEN_FILL
        else:
            ws1.cell(row=r, column=3).fill = RED_FILL
    if 'Realized' in label:
        ws1.cell(row=r, column=3).fill = RED_FILL

auto_width(ws1, max_w=30)

# ============================================================
# Sheets 2-6: AI ANALYSIS per stock
# ============================================================
def create_analysis_sheet(wb, sym, is_watchlist=False):
    title = f"AI — {sym}"
    ws = wb.create_sheet(title=title)

    d = daily.get(sym, {})
    l = llm.get(sym, {})
    j = judge.get(sym, {})

    ltp = fv(d.get('ltp'))

    # Title
    ws.merge_cells('A1:F1')
    label = 'WATCHLIST' if is_watchlist else 'PORTFOLIO HOLDING'
    ws['A1'] = f'{sym} — AI Analysis ({DATE}) [{label}]'
    ws['A1'].font = Font(bold=True, size=16, color='1F4E79')

    # Stock-specific notes
    note_row = 2
    if sym == 'GP':
        ws.merge_cells('A2:F2')
        ws['A2'] = 'NOTE: Bought AFTER record date — NO dividend applies'
        ws['A2'].font = Font(size=11, bold=True, color='CC0000')
        note_row = 3
    elif sym == 'HWAWELLTEX':
        ws.merge_cells('A2:F2')
        ws['A2'] = 'NOTE: T+2 locked until Thursday March 12, 2026. Bought 1000@44.8'
        ws['A2'].font = Font(size=11, bold=True, color='CC0000')
        note_row = 3

    # If portfolio, show holding info
    if not is_watchlist and sym in portfolio:
        info = portfolio[sym]
        note_row += 1
        ws.merge_cells(f'A{note_row}:F{note_row}')
        pl = (ltp - info['avg']) * info['qty'] if ltp else 0
        ws.cell(row=note_row, column=1,
                value=f"Holding: {info['qty']} shares @ {info['avg']:.1f} avg | LTP: {ltp:.1f} | P/L: {pl:+,.0f} BDT")
        ws.cell(row=note_row, column=1).font = Font(size=11, bold=True)

    # ── Section: Technical Indicators ──
    row = note_row + 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value='TECHNICAL INDICATORS')
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = MED_BLUE
        ws.cell(row=row, column=c).font = WHITE_FONT_MED

    indicators = [
        ('LTP', f"{ltp:.1f}" if ltp else 'N/A'),
        ('RSI', sv(d.get('rsi'))),
        ('StochRSI', sv(d.get('stoch_rsi'))),
        ('MACD Line', sv(d.get('macd_line'))),
        ('MACD Signal', sv(d.get('macd_signal'))),
        ('MACD Hist', sv(d.get('macd_hist'))),
        ('MACD Status', sv(d.get('macd_status'))),
        ('BB%', sv(d.get('bb_pct'))),
        ('ATR', sv(d.get('atr'))),
        ('ATR%', sv(d.get('atr_pct'))),
        ('Volatility', sv(d.get('volatility'))),
        ('Max Drawdown', sv(d.get('max_dd'))),
        ('Support', sv(d.get('support'))),
        ('Resistance', sv(d.get('resistance'))),
        ('50d Trend', sv(d.get('trend_50d'))),
        ('Avg Volume', sv(d.get('avg_vol'))),
        ('Vol Ratio', sv(d.get('vol_ratio'))),
        ('Score', sv(d.get('score'))),
    ]

    row += 1
    for i, (label, val) in enumerate(indicators):
        r = row + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=2).font = NORMAL_FONT
        alt = i % 2 == 1
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if alt:
                ws.cell(row=r, column=c).fill = LIGHT_GRAY

    # ── Section: AI Signal ──
    row = row + len(indicators) + 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value='AI SIGNAL (daily_analysis)')
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = MED_BLUE
        ws.cell(row=row, column=c).font = WHITE_FONT_MED

    ai_fields = [
        ('Action', sv(d.get('action'))),
        ('Entry Range', f"{sv(d.get('entry_low'))} - {sv(d.get('entry_high'))}"),
        ('Stop Loss', sv(d.get('sl'))),
        ('Target 1', sv(d.get('t1'))),
        ('Target 2', sv(d.get('t2'))),
        ('Risk %', sv(d.get('risk_pct'))),
        ('Reward %', sv(d.get('reward_pct'))),
        ('Wait Days', sv(d.get('wait_days'))),
        ('Vol Entry', sv(d.get('vol_entry'))),
    ]

    row += 1
    for i, (label, val) in enumerate(ai_fields):
        r = row + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=2).font = NORMAL_FONT
        alt = i % 2 == 1
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if alt:
                ws.cell(row=r, column=c).fill = LIGHT_GRAY

    row = row + len(ai_fields) + 1

    # ── Section: AI Reasoning ──
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value='REASONING (daily_analysis)')
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = MED_BLUE
        ws.cell(row=row, column=c).font = WHITE_FONT_MED

    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value=sv(d.get('reasoning')))
    ws.cell(row=row, column=1).alignment = WRAP
    ws.cell(row=row, column=1).font = NORMAL_FONT
    ws.row_dimensions[row].height = 80

    # ── Section: LLM Analysis ──
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value='LLM ANALYSIS (GPT)')
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = MED_BLUE
        ws.cell(row=row, column=c).font = WHITE_FONT_MED

    llm_fields = [
        ('Action', 'action'),
        ('Confidence', 'confidence'),
        ('Entry Range', None),  # special
        ('Stop Loss', 'sl'),
        ('Target 1', 't1'),
        ('Target 2', 't2'),
        ('Score', 'score'),
    ]

    row += 1
    for i, (label, key) in enumerate(llm_fields):
        r = row + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = BOLD_FONT
        if key is None:
            val = f"{sv(l.get('entry_low'))} - {sv(l.get('entry_high'))}"
        else:
            val = sv(l.get(key))
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=2).font = NORMAL_FONT
        alt = i % 2 == 1
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if alt:
                ws.cell(row=r, column=c).fill = LIGHT_GRAY

    row = row + len(llm_fields) + 1

    # LLM text fields
    llm_text_fields = [
        ('Reasoning', 'reasoning'),
        ('How to Buy', 'how_to_buy'),
        ('Volume Rule', 'volume_rule'),
        ('Next Day Plan', 'next_day_plan'),
        ('Sell Plan', 'sell_plan'),
    ]

    for label, key in llm_text_fields:
        val = sv(l.get(key))
        if not val:
            continue
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = LIGHT_BLUE
            ws.cell(row=row, column=c).border = THIN_BORDER

        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value=val)
        ws.cell(row=row, column=1).alignment = WRAP
        ws.cell(row=row, column=1).font = NORMAL_FONT
        ws.row_dimensions[row].height = 80
        row += 1

    # ── Section: Judge Analysis ──
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value='JUDGE ANALYSIS (Final Verdict)')
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = DARK_BLUE
        ws.cell(row=row, column=c).font = WHITE_FONT

    judge_kv = [
        ('Final Action', 'final_action'),
        ('Agreement', 'agreement'),
        ('Confidence', 'confidence'),
        ('Score Override', 'score'),
    ]

    row += 1
    for i, (label, key) in enumerate(judge_kv):
        r = row + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=2, value=sv(j.get(key)))
        ws.cell(row=r, column=2).font = NORMAL_FONT
        alt = i % 2 == 1
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if alt:
                ws.cell(row=r, column=c).fill = LIGHT_GRAY

    row = row + len(judge_kv) + 1

    # Judge text fields
    judge_text = [
        ('Reasoning', 'reasoning'),
        ('Key Risk', 'key_risk'),
    ]

    for label, key in judge_text:
        val = sv(j.get(key))
        if not val:
            continue
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = LIGHT_BLUE
            ws.cell(row=row, column=c).border = THIN_BORDER

        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value=val)
        ws.cell(row=row, column=1).alignment = WRAP
        ws.cell(row=row, column=1).font = NORMAL_FONT
        ws.row_dimensions[row].height = 80
        row += 1

    auto_width(ws, max_w=60)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 50

# Create sheets for portfolio stocks
for sym in portfolio:
    create_analysis_sheet(wb, sym, is_watchlist=False)

# Create sheet for watchlist
for sym in watchlist:
    create_analysis_sheet(wb, sym, is_watchlist=True)

# ============================================================
# Sheet 7: TOP BUY SIGNALS
# ============================================================
ws7 = wb.create_sheet(title="TOP BUY SIGNALS")

ws7.merge_cells('A1:K1')
ws7['A1'] = f'Top 20 Buy Signals — {DATE}'
ws7['A1'].font = Font(bold=True, size=16, color='1F4E79')

headers7 = ['Rank', 'Symbol', 'Action', 'Score', 'LTP', 'Entry Range', 'SL', 'T1', 'T2', 'MACD', 'RSI']
row = 3
for i, h in enumerate(headers7, 1):
    ws7.cell(row=row, column=i, value=h)
style_header_row(ws7, row, len(headers7))

row = 4
for idx, sig in enumerate(top_buys, 1):
    ws7.cell(row=row, column=1, value=idx)
    ws7.cell(row=row, column=2, value=sv(sig.get('symbol')))
    ws7.cell(row=row, column=3, value=sv(sig.get('action')))
    ws7.cell(row=row, column=4, value=fv(sig.get('score')))
    ws7.cell(row=row, column=5, value=round(fv(sig.get('ltp')), 1) if sig.get('ltp') else 'N/A')

    el = fv(sig.get('entry_low'))
    eh = fv(sig.get('entry_high'))
    if el and eh:
        ws7.cell(row=row, column=6, value=f"{el:.1f} - {eh:.1f}")
    else:
        ws7.cell(row=row, column=6, value='N/A')

    ws7.cell(row=row, column=7, value=round(fv(sig.get('sl')), 1) if sig.get('sl') else 'N/A')
    ws7.cell(row=row, column=8, value=round(fv(sig.get('t1')), 1) if sig.get('t1') else 'N/A')
    ws7.cell(row=row, column=9, value=round(fv(sig.get('t2')), 1) if sig.get('t2') else 'N/A')
    ws7.cell(row=row, column=10, value=sv(sig.get('macd_status')))
    ws7.cell(row=row, column=11, value=fv(sig.get('rsi')))

    alt = idx % 2 == 0
    style_data_row(ws7, row, len(headers7), alt=alt)
    row += 1

# Add how_to_buy for top 5
if top5_how:
    row += 2
    ws7.merge_cells(f'A{row}:K{row}')
    ws7.cell(row=row, column=1, value='HOW TO BUY — Top 5 Signals')
    ws7.cell(row=row, column=1).font = Font(bold=True, size=13, color='1F4E79')
    row += 1

    for sym, htb in top5_how.items():
        ws7.merge_cells(f'A{row}:K{row}')
        ws7.cell(row=row, column=1, value=sym)
        ws7.cell(row=row, column=1).font = BOLD_FONT
        for c in range(1, 12):
            ws7.cell(row=row, column=c).fill = LIGHT_BLUE
            ws7.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        ws7.merge_cells(f'A{row}:K{row}')
        ws7.cell(row=row, column=1, value=htb)
        ws7.cell(row=row, column=1).alignment = WRAP
        ws7.cell(row=row, column=1).font = NORMAL_FONT
        ws7.row_dimensions[row].height = 80
        row += 1

auto_width(ws7, max_w=35)

# ── Save ──
outpath = '/Users/shariarsourav/Desktop/DSE_FullAnalysis_10Mar2026.xlsx'
wb.save(outpath)
print(f"Saved: {outpath}")
print(f"Sheets: {wb.sheetnames}")
print(f"Portfolio stocks: {list(portfolio.keys())}")
print(f"Total invested: {total_invested:,.0f} | Current: {total_current:,.0f} | Unrealized: {unrealized_pl:+,.0f} | Net: {net_pl:+,.0f}")
