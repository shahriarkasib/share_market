"""Generate comprehensive DSE Historical Analysis Excel report.

Uses 10+ years of daily_prices data (Dec 2014 - Mar 2026, ~590K rows)
to produce sector/stock seasonality, indicator floor analysis, and
capitulation radar.

Output: /Users/shariarsourav/Desktop/DSE_Historical_Analysis.xlsx
"""

import sys
import time
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)

# ── Constants ──────────────────────────────────────────────
DATABASE_URL = (
    "postgresql://postgres.iihlezpkpllacztoaguc:"
    "160021062Ss%23%23@aws-1-ap-southeast-1.pooler.supabase.com:6543/postgres"
)
OUTPUT_PATH = "/Users/shariarsourav/Desktop/DSE_Historical_Analysis.xlsx"

MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

# ── Styles ─────────────────────────────────────────────────
DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
MED_BLUE = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4CC", end_color="FCE4CC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
DARK_RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
WHITE_FONT_MED = Font(color="FFFFFF", bold=True, size=10)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header_row(ws, row, cols, fill=None, font=None):
    fill = fill or DARK_BLUE
    font = font or WHITE_FONT
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, cols, alt=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.font = NORMAL_FONT
        if alt:
            cell.fill = LIGHT_GRAY


def auto_width(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                max_len = max(max_len, longest)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ── DB connection ──────────────────────────────────────────
def get_conn():
    return psycopg2.connect(
        DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor
    )


# ── Data loading ───────────────────────────────────────────
def load_data():
    """Load all A-category daily prices with sector info."""
    print("Loading data from PostgreSQL...")
    t0 = time.time()
    conn = get_conn()

    # Get A-category stocks with sectors
    cur = conn.cursor()
    cur.execute(
        "SELECT symbol, sector FROM fundamentals WHERE category = 'A' AND sector IS NOT NULL"
    )
    stock_sectors = {r["symbol"]: r["sector"] for r in cur.fetchall()}
    print(f"  A-category stocks: {len(stock_sectors)}")

    # Load all daily prices for these symbols
    symbols = tuple(stock_sectors.keys())
    cur.execute(
        """
        SELECT dp.symbol, dp.date, dp.open, dp.high, dp.low, dp.close, dp.volume
        FROM daily_prices dp
        WHERE dp.symbol IN %s AND dp.close > 0
        ORDER BY dp.symbol, dp.date
        """,
        (symbols,),
    )
    rows = cur.fetchall()
    conn.close()

    df = pd.DataFrame(rows)
    for col in ["open", "high", "low", "close", "volume"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["date"] = pd.to_datetime(df["date"])
    df["sector"] = df["symbol"].map(stock_sectors)

    elapsed = time.time() - t0
    print(f"  Loaded {len(df):,} rows in {elapsed:.1f}s")
    print(f"  Date range: {df['date'].min().date()} to {df['date'].max().date()}")
    return df, stock_sectors


# ── Sheet 1: Sector Seasonality ────────────────────────────
def compute_sector_seasonality(df):
    """Compute monthly returns grouped by sector."""
    print("\nComputing sector seasonality...")
    t0 = time.time()

    df = df.copy()
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month

    # Monthly returns per stock: last close / first close - 1
    monthly_returns = []
    for (symbol, year, month), grp in df.groupby(["symbol", "year", "month"]):
        if len(grp) < 2:
            continue
        first_close = grp.iloc[0]["close"]
        last_close = grp.iloc[-1]["close"]
        if first_close > 0:
            ret = (last_close / first_close - 1) * 100
            monthly_returns.append({
                "symbol": symbol,
                "sector": grp.iloc[0]["sector"],
                "year": year,
                "month": month,
                "return_pct": ret,
            })

    mr_df = pd.DataFrame(monthly_returns)

    # Aggregate by sector + month
    results = []
    for (sector, month), grp in mr_df.groupby(["sector", "month"]):
        # Top 3 best-performing stocks in this sector-month
        stock_avg = grp.groupby("symbol")["return_pct"].mean()
        top3 = stock_avg.nlargest(3).index.tolist()

        results.append({
            "sector": sector,
            "month": month,
            "month_name": MONTH_NAMES[month - 1],
            "avg_return_pct": round(grp["return_pct"].mean(), 2),
            "median_return_pct": round(grp["return_pct"].median(), 2),
            "win_rate_pct": round((grp["return_pct"] > 0).mean() * 100, 1),
            "sample_count": len(grp),
            "best_stocks": ", ".join(top3),
        })

    result_df = pd.DataFrame(results).sort_values(["sector", "month"])
    print(f"  {len(result_df)} sector-month combos in {time.time() - t0:.1f}s")
    return result_df


# ── Sheet 2: Stock Seasonality ─────────────────────────────
def compute_stock_seasonality(df):
    """Compute monthly returns per individual stock."""
    print("\nComputing stock seasonality...")
    t0 = time.time()

    df = df.copy()
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month

    monthly_returns = []
    for (symbol, year, month), grp in df.groupby(["symbol", "year", "month"]):
        if len(grp) < 2:
            continue
        first_close = grp.iloc[0]["close"]
        last_close = grp.iloc[-1]["close"]
        if first_close > 0:
            ret = (last_close / first_close - 1) * 100
            monthly_returns.append({
                "symbol": symbol,
                "sector": grp.iloc[0]["sector"],
                "year": year,
                "month": month,
                "return_pct": ret,
            })

    mr_df = pd.DataFrame(monthly_returns)

    results = []
    for (symbol, month), grp in mr_df.groupby(["symbol", "month"]):
        results.append({
            "symbol": symbol,
            "sector": grp.iloc[0]["sector"],
            "month_name": MONTH_NAMES[month - 1],
            "month": month,
            "avg_return_pct": round(grp["return_pct"].mean(), 2),
            "median_return_pct": round(grp["return_pct"].median(), 2),
            "win_rate_pct": round((grp["return_pct"] > 0).mean() * 100, 1),
            "sample_years": len(grp),
        })

    result_df = pd.DataFrame(results).sort_values(["symbol", "month"])
    print(f"  {len(result_df)} stock-month combos in {time.time() - t0:.1f}s")
    return result_df


# ── Indicator computation (vectorized) ─────────────────────
def compute_rsi(close, period=14):
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = (-delta).where(delta < 0, 0.0)
    avg_gain = gain.rolling(window=period, min_periods=period).mean()
    avg_loss = loss.rolling(window=period, min_periods=period).mean()
    for i in range(period, len(close)):
        avg_gain.iloc[i] = (avg_gain.iloc[i - 1] * (period - 1) + gain.iloc[i]) / period
        avg_loss.iloc[i] = (avg_loss.iloc[i - 1] * (period - 1) + loss.iloc[i]) / period
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def compute_indicators_for_stock(stock_df):
    """Compute 7 indicators for a single stock DataFrame.

    Returns DataFrame with: rsi, macd_hist, stochrsi, mfi, williams_r, cmf, bb_pct
    """
    df = stock_df.copy().reset_index(drop=True)
    if len(df) < 30:
        return None

    close = df["close"]
    high = df["high"]
    low = df["low"]
    volume = df["volume"]

    # RSI(14)
    df["rsi"] = compute_rsi(close, 14)

    # MACD histogram
    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    macd = ema12 - ema26
    signal = macd.ewm(span=9, adjust=False).mean()
    df["macd_hist"] = macd - signal

    # StochRSI
    rsi = df["rsi"]
    rsi_low = rsi.rolling(14).min()
    rsi_high = rsi.rolling(14).max()
    rsi_range = (rsi_high - rsi_low).replace(0, np.nan)
    stochrsi_raw = ((rsi - rsi_low) / rsi_range) * 100
    df["stochrsi"] = stochrsi_raw.rolling(3).mean()

    # MFI(14)
    typical = (high + low + close) / 3
    raw_mf = typical * volume
    pos = raw_mf.where(typical > typical.shift(1), 0.0)
    neg = raw_mf.where(typical < typical.shift(1), 0.0)
    pos_sum = pos.rolling(14).sum()
    neg_sum = neg.rolling(14).sum()
    df["mfi"] = 100 - (100 / (1 + pos_sum / neg_sum.replace(0, np.nan)))

    # Williams %R(14)
    high14 = high.rolling(14).max()
    low14 = low.rolling(14).min()
    price_range = (high14 - low14).replace(0, np.nan)
    df["williams_r"] = ((high14 - close) / price_range) * -100

    # CMF(20)
    clv = ((close - low) - (high - close)) / (high - low).replace(0, np.nan)
    mfv = clv * volume
    df["cmf"] = mfv.rolling(20).sum() / volume.rolling(20).sum()

    # BB%(20,2)
    sma20 = close.rolling(20).mean()
    std20 = close.rolling(20).std()
    bb_upper = sma20 + 2 * std20
    bb_lower = sma20 - 2 * std20
    bb_range = (bb_upper - bb_lower).replace(0, np.nan)
    df["bb_pct"] = (close - bb_lower) / bb_range

    return df


# ── Sheet 3: Indicator Floors ──────────────────────────────
INDICATOR_COLS = ["rsi", "macd_hist", "stochrsi", "mfi", "williams_r", "cmf", "bb_pct"]

def classify_trajectory(current, floor, ceiling, recent_values):
    """Classify trajectory relative to floor.

    Returns one of: AT_FLOOR, FALLING_TOWARD, BOUNCING_UP, RISING_AWAY, NEUTRAL
    """
    if ceiling == floor:
        return "NEUTRAL"
    distance_pct = (current - floor) / (ceiling - floor) * 100

    if distance_pct < 10:
        return "AT_FLOOR"

    # Check trend from last 5 values
    if len(recent_values) >= 5:
        recent = recent_values[-5:]
        trend = recent[-1] - recent[0]
        if distance_pct < 30 and trend < 0:
            return "FALLING_TOWARD"
        # Was near floor recently (within last 10 days) and now rising
        if len(recent_values) >= 10:
            recent_min_dist = min(
                (v - floor) / (ceiling - floor) * 100
                for v in recent_values[-10:]
                if not np.isnan(v)
            ) if any(not np.isnan(v) for v in recent_values[-10:]) else 50
            if recent_min_dist < 15 and trend > 0:
                return "BOUNCING_UP"
        if trend > 0 and distance_pct > 30:
            return "RISING_AWAY"

    return "NEUTRAL"


def compute_indicator_floors(df, stock_sectors):
    """Compute indicator floor analysis for all A-category stocks."""
    print("\nComputing indicator floors (this takes a few minutes)...")
    t0 = time.time()

    results = []
    symbols = df["symbol"].unique()
    total = len(symbols)

    for idx, symbol in enumerate(symbols):
        if (idx + 1) % 50 == 0:
            print(f"  Processing {idx + 1}/{total}...")

        stock_df = df[df["symbol"] == symbol].copy()
        enriched = compute_indicators_for_stock(stock_df)
        if enriched is None:
            continue

        sector = stock_sectors.get(symbol, "Unknown")

        for ind in INDICATOR_COLS:
            series = enriched[ind].dropna()
            if len(series) < 50:
                continue

            floor_val = series.min()
            ceiling_val = series.max()
            floor_date = enriched.loc[series.idxmin(), "date"]
            current_val = series.iloc[-1]

            if ceiling_val == floor_val:
                distance_pct = 50.0
            else:
                distance_pct = (current_val - floor_val) / (ceiling_val - floor_val) * 100

            # Recent values for trajectory
            recent = series.tail(20).tolist()
            trajectory = classify_trajectory(current_val, floor_val, ceiling_val, recent)

            # Avg bounce: forward 10-day return when indicator in bottom 10th pctile
            threshold = series.quantile(0.10)
            bottom_mask = series <= threshold
            bottom_indices = series[bottom_mask].index.tolist()

            bounce_returns = []
            for bi in bottom_indices:
                future_idx = bi + 10
                if future_idx < len(enriched):
                    entry_price = enriched.loc[bi, "close"]
                    exit_price = enriched.loc[future_idx, "close"]
                    if entry_price > 0:
                        bounce_returns.append((exit_price / entry_price - 1) * 100)

            avg_bounce = round(np.mean(bounce_returns), 2) if bounce_returns else None

            results.append({
                "symbol": symbol,
                "sector": sector,
                "indicator": ind,
                "floor": round(floor_val, 4),
                "floor_date": floor_date.strftime("%Y-%m-%d") if hasattr(floor_date, "strftime") else str(floor_date),
                "ceiling": round(ceiling_val, 4),
                "current": round(current_val, 4),
                "distance_pct": round(distance_pct, 1),
                "trajectory": trajectory,
                "avg_bounce_pct": avg_bounce,
            })

    result_df = pd.DataFrame(results)

    # Compute capitulation_score per stock
    cap_scores = {}
    for symbol, grp in result_df.groupby("symbol"):
        score = sum(
            1 for _, r in grp.iterrows()
            if r["trajectory"] in ("AT_FLOOR", "FALLING_TOWARD")
        )
        cap_scores[symbol] = score

    result_df["capitulation_score"] = result_df["symbol"].map(cap_scores)

    elapsed = time.time() - t0
    print(f"  {len(result_df)} indicator-floor rows for {total} stocks in {elapsed:.1f}s")
    return result_df


# ── Write Excel ────────────────────────────────────────────
def write_excel(sector_season, stock_season, indicator_floors):
    """Write all sheets to the output Excel file."""
    print(f"\nWriting Excel to {OUTPUT_PATH}...")
    wb = Workbook()

    # ── Sheet 1: Sector Seasonality ──
    ws1 = wb.active
    ws1.title = "Sector Seasonality"
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "DSE Sector Seasonality (Dec 2014 - Mar 2026, A-Category)"
    ws1["A1"].font = Font(bold=True, size=14, color="1F4E79")

    headers = [
        "Sector", "Month", "Avg Return %", "Median Return %",
        "Win Rate %", "Sample Count", "Best Stocks",
    ]
    row = 3
    for i, h in enumerate(headers, 1):
        ws1.cell(row=row, column=i, value=h)
    style_header_row(ws1, row, len(headers))

    row = 4
    prev_sector = None
    for idx, (_, r) in enumerate(sector_season.iterrows()):
        ws1.cell(row=row, column=1, value=r["sector"])
        ws1.cell(row=row, column=2, value=r["month_name"])
        ws1.cell(row=row, column=3, value=r["avg_return_pct"])
        ws1.cell(row=row, column=4, value=r["median_return_pct"])
        ws1.cell(row=row, column=5, value=r["win_rate_pct"])
        ws1.cell(row=row, column=6, value=r["sample_count"])
        ws1.cell(row=row, column=7, value=r["best_stocks"])

        alt = r["sector"] != prev_sector and prev_sector is not None
        if alt:
            prev_sector = r["sector"]
        if prev_sector is None:
            prev_sector = r["sector"]
        style_data_row(ws1, row, len(headers), alt=(idx % 2 == 1))

        # Color avg return
        avg_cell = ws1.cell(row=row, column=3)
        if r["avg_return_pct"] > 2:
            avg_cell.fill = GREEN_FILL
        elif r["avg_return_pct"] < -2:
            avg_cell.fill = RED_FILL

        # Color win rate
        wr_cell = ws1.cell(row=row, column=5)
        if r["win_rate_pct"] > 60:
            wr_cell.fill = GREEN_FILL
        elif r["win_rate_pct"] < 40:
            wr_cell.fill = RED_FILL

        row += 1

    auto_width(ws1)

    # ── Sheet 2: Stock Seasonality ──
    ws2 = wb.create_sheet("Stock Seasonality")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "DSE Stock-Level Seasonality (Dec 2014 - Mar 2026, A-Category)"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")

    headers2 = [
        "Symbol", "Sector", "Month", "Avg Return %",
        "Median Return %", "Win Rate %", "Sample Years",
    ]
    row = 3
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=row, column=i, value=h)
    style_header_row(ws2, row, len(headers2))

    row = 4
    for idx, (_, r) in enumerate(stock_season.iterrows()):
        ws2.cell(row=row, column=1, value=r["symbol"])
        ws2.cell(row=row, column=2, value=r["sector"])
        ws2.cell(row=row, column=3, value=r["month_name"])
        ws2.cell(row=row, column=4, value=r["avg_return_pct"])
        ws2.cell(row=row, column=5, value=r["median_return_pct"])
        ws2.cell(row=row, column=6, value=r["win_rate_pct"])
        ws2.cell(row=row, column=7, value=r["sample_years"])

        style_data_row(ws2, row, len(headers2), alt=(idx % 2 == 1))

        avg_cell = ws2.cell(row=row, column=4)
        if r["avg_return_pct"] > 3:
            avg_cell.fill = GREEN_FILL
        elif r["avg_return_pct"] < -3:
            avg_cell.fill = RED_FILL

        row += 1

    auto_width(ws2)

    # ── Sheet 3: Indicator Floors ──
    ws3 = wb.create_sheet("Indicator Floors")
    ws3.merge_cells("A1:K1")
    ws3["A1"] = "DSE Indicator Floor Analysis (Capitulation Detection)"
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E79")

    headers3 = [
        "Symbol", "Sector", "Indicator", "Floor", "Floor Date",
        "Ceiling", "Current", "Distance %", "Trajectory",
        "Avg Bounce %", "Cap. Score",
    ]
    row = 3
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=row, column=i, value=h)
    style_header_row(ws3, row, len(headers3))

    # Sort by capitulation_score desc, symbol
    floors_sorted = indicator_floors.sort_values(
        ["capitulation_score", "symbol", "indicator"],
        ascending=[False, True, True],
    )

    row = 4
    for idx, (_, r) in enumerate(floors_sorted.iterrows()):
        ws3.cell(row=row, column=1, value=r["symbol"])
        ws3.cell(row=row, column=2, value=r["sector"])
        ws3.cell(row=row, column=3, value=r["indicator"])
        ws3.cell(row=row, column=4, value=r["floor"])
        ws3.cell(row=row, column=5, value=r["floor_date"])
        ws3.cell(row=row, column=6, value=r["ceiling"])
        ws3.cell(row=row, column=7, value=r["current"])
        ws3.cell(row=row, column=8, value=r["distance_pct"])
        ws3.cell(row=row, column=9, value=r["trajectory"])
        ws3.cell(row=row, column=10, value=r["avg_bounce_pct"])
        ws3.cell(row=row, column=11, value=r["capitulation_score"])

        style_data_row(ws3, row, len(headers3), alt=(idx % 2 == 1))

        # Color trajectory
        traj_cell = ws3.cell(row=row, column=9)
        traj = r["trajectory"]
        if traj == "AT_FLOOR":
            traj_cell.fill = DARK_RED_FILL
            traj_cell.font = Font(color="FFFFFF", bold=True, size=10)
        elif traj == "FALLING_TOWARD":
            traj_cell.fill = RED_FILL
        elif traj == "BOUNCING_UP":
            traj_cell.fill = GREEN_FILL
        elif traj == "RISING_AWAY":
            traj_cell.fill = LIGHT_BLUE

        # Color cap score
        cap_cell = ws3.cell(row=row, column=11)
        score = r["capitulation_score"]
        if score >= 5:
            cap_cell.fill = DARK_RED_FILL
            cap_cell.font = Font(color="FFFFFF", bold=True, size=10)
        elif score >= 3:
            cap_cell.fill = RED_FILL
        elif score >= 2:
            cap_cell.fill = ORANGE_FILL

        row += 1

    auto_width(ws3)

    # ── Sheet 4: Capitulation Radar ──
    ws4 = wb.create_sheet("Capitulation Radar")
    ws4.merge_cells("A1:I1")
    ws4["A1"] = "Top 30 Capitulation Candidates (Most Indicators Near Historical Floors)"
    ws4["A1"].font = Font(bold=True, size=14, color="1F4E79")

    # Get unique stocks sorted by cap score
    stock_cap = (
        indicator_floors.groupby(["symbol", "sector"])
        .agg(
            capitulation_score=("capitulation_score", "first"),
            at_floor_indicators=("trajectory", lambda x: ", ".join(
                r["indicator"] for _, r in indicator_floors.loc[x.index].iterrows()
                if r["trajectory"] in ("AT_FLOOR", "FALLING_TOWARD")
            )),
            avg_distance=("distance_pct", "mean"),
            avg_bounce=("avg_bounce_pct", "mean"),
        )
        .reset_index()
        .sort_values("capitulation_score", ascending=False)
        .head(30)
    )

    headers4 = [
        "Rank", "Symbol", "Sector", "Cap. Score (0-7)",
        "Indicators at/near Floor", "Avg Distance %",
        "Avg Bounce %", "Signal",
    ]
    row = 3
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=row, column=i, value=h)
    style_header_row(ws4, row, len(headers4))

    row = 4
    for idx, (_, r) in enumerate(stock_cap.iterrows()):
        score = r["capitulation_score"]
        if score >= 5:
            signal = "STRONG CAPITULATION"
        elif score >= 3:
            signal = "MODERATE CAPITULATION"
        elif score >= 2:
            signal = "MILD STRESS"
        else:
            signal = "NEUTRAL"

        ws4.cell(row=row, column=1, value=idx + 1)
        ws4.cell(row=row, column=2, value=r["symbol"])
        ws4.cell(row=row, column=3, value=r["sector"])
        ws4.cell(row=row, column=4, value=score)
        ws4.cell(row=row, column=5, value=r["at_floor_indicators"])
        ws4.cell(row=row, column=6, value=round(r["avg_distance"], 1))
        ws4.cell(row=row, column=7, value=round(r["avg_bounce"], 2) if pd.notna(r["avg_bounce"]) else "N/A")
        ws4.cell(row=row, column=8, value=signal)

        style_data_row(ws4, row, len(headers4), alt=(idx % 2 == 1))

        # Color signal
        sig_cell = ws4.cell(row=row, column=8)
        if "STRONG" in signal:
            sig_cell.fill = DARK_RED_FILL
            sig_cell.font = Font(color="FFFFFF", bold=True, size=10)
        elif "MODERATE" in signal:
            sig_cell.fill = RED_FILL
        elif "MILD" in signal:
            sig_cell.fill = ORANGE_FILL

        # Color cap score
        cap_cell = ws4.cell(row=row, column=4)
        if score >= 5:
            cap_cell.fill = DARK_RED_FILL
            cap_cell.font = Font(color="FFFFFF", bold=True, size=10)
        elif score >= 3:
            cap_cell.fill = RED_FILL
        elif score >= 2:
            cap_cell.fill = ORANGE_FILL

        row += 1

    auto_width(ws4)
    ws4.column_dimensions["E"].width = 40

    # ── Sheet 5: Sector Capitulation ──
    ws5 = wb.create_sheet("Sector Capitulation")
    ws5.merge_cells("A1:G1")
    ws5["A1"] = "Sector-Level Capitulation Summary"
    ws5["A1"].font = Font(bold=True, size=14, color="1F4E79")

    sector_cap = (
        indicator_floors.groupby("sector")
        .agg(
            avg_cap_score=("capitulation_score", "mean"),
            max_cap_score=("capitulation_score", "max"),
            stocks_at_floor=("trajectory", lambda x: sum(1 for t in x if t == "AT_FLOOR")),
            stocks_falling=("trajectory", lambda x: sum(1 for t in x if t == "FALLING_TOWARD")),
            stocks_bouncing=("trajectory", lambda x: sum(1 for t in x if t == "BOUNCING_UP")),
            avg_distance=("distance_pct", "mean"),
            total_indicators=("indicator", "count"),
        )
        .reset_index()
        .sort_values("avg_cap_score", ascending=False)
    )

    headers5 = [
        "Sector", "Avg Cap Score", "Max Cap Score",
        "# At Floor", "# Falling", "# Bouncing",
        "Avg Distance %",
    ]
    row = 3
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=row, column=i, value=h)
    style_header_row(ws5, row, len(headers5))

    row = 4
    for idx, (_, r) in enumerate(sector_cap.iterrows()):
        ws5.cell(row=row, column=1, value=r["sector"])
        ws5.cell(row=row, column=2, value=round(r["avg_cap_score"], 1))
        ws5.cell(row=row, column=3, value=int(r["max_cap_score"]))
        ws5.cell(row=row, column=4, value=int(r["stocks_at_floor"]))
        ws5.cell(row=row, column=5, value=int(r["stocks_falling"]))
        ws5.cell(row=row, column=6, value=int(r["stocks_bouncing"]))
        ws5.cell(row=row, column=7, value=round(r["avg_distance"], 1))

        style_data_row(ws5, row, len(headers5), alt=(idx % 2 == 1))

        # Color avg cap score
        cap_cell = ws5.cell(row=row, column=2)
        if r["avg_cap_score"] >= 3:
            cap_cell.fill = RED_FILL
        elif r["avg_cap_score"] >= 2:
            cap_cell.fill = ORANGE_FILL

        row += 1

    auto_width(ws5)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


# ── Main ───────────────────────────────────────────────────
def main():
    t_start = time.time()
    print("=" * 60)
    print("DSE Historical Analysis Excel Generator")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    df, stock_sectors = load_data()

    sector_season = compute_sector_seasonality(df)
    stock_season = compute_stock_seasonality(df)
    indicator_floors = compute_indicator_floors(df, stock_sectors)

    write_excel(sector_season, stock_season, indicator_floors)

    elapsed = time.time() - t_start
    print(f"\nTotal time: {elapsed:.1f}s")


if __name__ == "__main__":
    main()
