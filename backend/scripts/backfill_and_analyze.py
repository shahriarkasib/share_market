#!/usr/bin/env python3
"""
Backfill daily_prices (Mar 2024 → Mar 2025), compute indicators, generate Excel.

Usage:
    cd /Users/shariarsourav/Desktop/share_market/backend
    source venv/bin/activate
    python scripts/backfill_and_analyze.py
"""

import sys
import os
import time
import warnings
import functools
import numpy as np
import pandas as pd
from datetime import datetime

# Force unbuffered output
print = functools.partial(print, flush=True)

warnings.filterwarnings("ignore")

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import get_connection

OUTPUT_PATH = "/Users/shariarsourav/Desktop/DSE_Historical_Analysis.xlsx"


# ═══════════════════════════════════════════════════════════════════════
# PART 1: Backfill daily_prices from bdshare
# ═══════════════════════════════════════════════════════════════════════

def backfill_prices():
    """Fetch Mar 2024 → Mar 2025 data from bdshare for A-category stocks."""
    from bdshare import get_historical_data

    conn = get_connection()
    rows = conn.execute(
        "SELECT symbol FROM fundamentals WHERE category = 'A' ORDER BY symbol"
    ).fetchall()
    symbols = [r["symbol"] for r in rows]
    conn.close()

    print(f"[BACKFILL] {len(symbols)} A-category symbols to backfill")

    total_inserted = 0
    failed = []

    for i, sym in enumerate(symbols):
        if (i + 1) % 10 == 0 or i == 0:
            print(f"  [{i+1}/{len(symbols)}] Processing {sym}...")

        try:
            df = get_historical_data(
                start="2024-03-01", end="2025-03-02", code=sym
            )
        except Exception:
            df = None

        if df is None or df.empty:
            failed.append(sym)
            time.sleep(0.3)
            continue

        conn = get_connection()
        inserted = 0
        for date_val, row in df.iterrows():
            # date_val is the index (date string like '2024-03-18')
            close_price = float(row.get("close", 0) or row.get("ltp", 0))
            open_price = float(row.get("open", 0) or close_price)
            high_price = float(row.get("high", 0) or close_price)
            low_price = float(row.get("low", 0) or close_price)
            volume = int(row.get("volume", 0) or 0)

            if close_price <= 0:
                continue

            # Round to DSE tick size
            close_price = round(close_price, 1)
            open_price = round(open_price, 1)
            high_price = round(high_price, 1)
            low_price = round(low_price, 1)

            date_str = str(date_val)[:10]

            try:
                conn.execute(
                    """INSERT INTO daily_prices (symbol, date, open, high, low, close, volume)
                       VALUES (%s, %s, %s, %s, %s, %s, %s)
                       ON CONFLICT (symbol, date) DO NOTHING""",
                    (sym, date_str, open_price, high_price, low_price, close_price, volume),
                )
                inserted += 1
            except Exception:
                pass

        conn.commit()
        conn.close()
        total_inserted += inserted
        time.sleep(0.5)

    print(f"[BACKFILL] Done. Inserted {total_inserted} rows. Failed/empty: {len(failed)} stocks")
    if failed:
        print(f"  Failed symbols: {', '.join(failed[:20])}{'...' if len(failed) > 20 else ''}")
    return symbols


# ═══════════════════════════════════════════════════════════════════════
# PART 2: Compute technical indicators from raw prices
# ═══════════════════════════════════════════════════════════════════════

def compute_ema(series, period):
    """Exponential moving average."""
    return series.ewm(span=period, adjust=False).mean()


def compute_rsi(close, period=14):
    """RSI using exponential moving average of gains/losses."""
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    avg_gain = gain.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi


def compute_macd(close, fast=12, slow=26, signal=9):
    """MACD line, signal, histogram."""
    ema_fast = compute_ema(close, fast)
    ema_slow = compute_ema(close, slow)
    macd_line = ema_fast - ema_slow
    macd_signal = compute_ema(macd_line, signal)
    macd_hist = macd_line - macd_signal
    return macd_line, macd_signal, macd_hist


def compute_stoch_rsi(close, period=14):
    """Stochastic RSI: (RSI - RSI_low) / (RSI_high - RSI_low) * 100."""
    rsi = compute_rsi(close, period)
    rsi_low = rsi.rolling(period).min()
    rsi_high = rsi.rolling(period).max()
    denom = (rsi_high - rsi_low).replace(0, np.nan)
    stoch_rsi = (rsi - rsi_low) / denom * 100
    return stoch_rsi


def compute_mfi(high, low, close, volume, period=14):
    """Money Flow Index."""
    typical_price = (high + low + close) / 3
    raw_mf = typical_price * volume
    delta = typical_price.diff()
    pos_mf = pd.Series(np.where(delta > 0, raw_mf, 0), index=close.index)
    neg_mf = pd.Series(np.where(delta < 0, raw_mf, 0), index=close.index)
    pos_sum = pos_mf.rolling(period).sum()
    neg_sum = neg_mf.rolling(period).sum().replace(0, np.nan)
    mfi = 100 - (100 / (1 + pos_sum / neg_sum))
    return mfi


def compute_williams_r(high, low, close, period=14):
    """Williams %R."""
    hh = high.rolling(period).max()
    ll = low.rolling(period).min()
    denom = (hh - ll).replace(0, np.nan)
    wr = (hh - close) / denom * -100
    return wr


def compute_bb_pct(close, period=20, std_mult=2):
    """Bollinger Band %B."""
    sma = close.rolling(period).mean()
    std = close.rolling(period).std()
    upper = sma + std_mult * std
    lower = sma - std_mult * std
    denom = (upper - lower).replace(0, np.nan)
    bb_pct = (close - lower) / denom * 100
    return bb_pct


def compute_cmf(high, low, close, volume, period=20):
    """Chaikin Money Flow."""
    denom_hl = (high - low).replace(0, np.nan)
    mf_mult = ((close - low) - (high - close)) / denom_hl
    mf_volume = mf_mult * volume
    cmf = mf_volume.rolling(period).sum() / volume.rolling(period).sum().replace(0, np.nan)
    return cmf


def compute_adx(high, low, close, period=14):
    """Average Directional Index."""
    plus_dm = high.diff()
    minus_dm = -low.diff()
    plus_dm = pd.Series(
        np.where((plus_dm > minus_dm) & (plus_dm > 0), plus_dm, 0),
        index=close.index
    )
    minus_dm = pd.Series(
        np.where((minus_dm > plus_dm) & (minus_dm > 0), minus_dm, 0),
        index=close.index
    )

    tr1 = high - low
    tr2 = (high - close.shift()).abs()
    tr3 = (low - close.shift()).abs()
    tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)

    atr = tr.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    plus_di = 100 * plus_dm.ewm(alpha=1/period, min_periods=period, adjust=False).mean() / atr.replace(0, np.nan)
    minus_di = 100 * minus_dm.ewm(alpha=1/period, min_periods=period, adjust=False).mean() / atr.replace(0, np.nan)

    dx_denom = (plus_di + minus_di).replace(0, np.nan)
    dx = (plus_di - minus_di).abs() / dx_denom * 100
    adx = dx.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    return adx


def compute_all_indicators(df):
    """Compute all indicators for a stock DataFrame.

    df must have columns: date, open, high, low, close, volume (sorted by date asc).
    Returns the DataFrame with indicator columns added.
    """
    c = df["close"].astype(float)
    h = df["high"].astype(float)
    l = df["low"].astype(float)
    v = df["volume"].astype(float)

    df = df.copy()
    df["rsi"] = compute_rsi(c)
    macd_line, macd_signal, macd_hist = compute_macd(c)
    df["macd_line"] = macd_line
    df["macd_signal"] = macd_signal
    df["macd_hist"] = macd_hist
    df["stoch_rsi"] = compute_stoch_rsi(c)
    df["mfi"] = compute_mfi(h, l, c, v)
    df["williams_r"] = compute_williams_r(h, l, c)
    df["cmf"] = compute_cmf(h, l, c, v)
    df["bb_pct"] = compute_bb_pct(c)
    df["adx"] = compute_adx(h, l, c)

    return df


# ═══════════════════════════════════════════════════════════════════════
# PART 3: Load all price data, compute indicators, build Excel
# ═══════════════════════════════════════════════════════════════════════

def load_all_prices():
    """Load all daily_prices from DB into a dict of DataFrames keyed by symbol."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT symbol, date, open, high, low, close, volume FROM daily_prices ORDER BY symbol, date"
    ).fetchall()
    conn.close()

    data = {}
    for r in rows:
        sym = r["symbol"]
        if sym not in data:
            data[sym] = []
        data[sym].append({
            "date": r["date"],
            "open": float(r["open"] or 0),
            "high": float(r["high"] or 0),
            "low": float(r["low"] or 0),
            "close": float(r["close"] or 0),
            "volume": int(r["volume"] or 0),
        })

    result = {}
    for sym, records in data.items():
        df = pd.DataFrame(records)
        df["date"] = pd.to_datetime(df["date"])
        df = df.sort_values("date").reset_index(drop=True)
        if len(df) >= 30:  # Need minimum data for indicators
            result[sym] = df

    print(f"[DATA] Loaded {len(result)} stocks with >= 30 days of data")
    return result


def load_sector_map():
    """Load symbol -> sector mapping."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT symbol, sector, category FROM fundamentals WHERE sector IS NOT NULL"
    ).fetchall()
    conn.close()
    sector_map = {r["symbol"]: r["sector"] for r in rows}
    category_map = {r["symbol"]: r["category"] for r in rows}
    return sector_map, category_map


def build_seasonality_sheets(all_prices, sector_map):
    """Build Sheet 1 (Sector Seasonality) and Sheet 2 (Stock Seasonality)."""
    print("[EXCEL] Computing seasonality...")

    stock_monthly = []

    for sym, df in all_prices.items():
        if sym not in sector_map:
            continue
        sector = sector_map[sym]
        df = df.copy()
        df["year_month"] = df["date"].dt.to_period("M")

        for ym, grp in df.groupby("year_month"):
            if len(grp) < 2:
                continue
            first_close = grp.iloc[0]["close"]
            last_close = grp.iloc[-1]["close"]
            if first_close <= 0:
                continue
            ret = (last_close / first_close - 1) * 100
            stock_monthly.append({
                "symbol": sym,
                "sector": sector,
                "month": ym.month,
                "month_name": ym.strftime("%b"),
                "year": ym.year,
                "return_pct": ret,
            })

    mdf = pd.DataFrame(stock_monthly)
    if mdf.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Sheet 2: Stock Seasonality
    stock_season = (
        mdf.groupby(["symbol", "sector", "month", "month_name"])
        .agg(
            avg_return_pct=("return_pct", "mean"),
            median_return_pct=("return_pct", "median"),
            win_rate_pct=("return_pct", lambda x: (x > 0).mean() * 100),
            sample_count=("return_pct", "count"),
        )
        .reset_index()
        .sort_values(["symbol", "month"])
    )

    # Sheet 1: Sector Seasonality
    sector_records = []
    for (sector, month, month_name), grp in mdf.groupby(["sector", "month", "month_name"]):
        # Per-stock avg returns for finding best stocks
        stock_avg = grp.groupby("symbol")["return_pct"].mean().sort_values(ascending=False)
        best_stocks = ", ".join(stock_avg.head(3).index.tolist())

        sector_records.append({
            "sector": sector,
            "month": month,
            "month_name": month_name,
            "avg_return_pct": round(grp["return_pct"].mean(), 2),
            "median_return_pct": round(grp["return_pct"].median(), 2),
            "win_rate_pct": round((grp["return_pct"] > 0).mean() * 100, 1),
            "sample_count": len(grp),
            "best_stocks": best_stocks,
        })

    sector_season = pd.DataFrame(sector_records).sort_values(["sector", "month"])

    # Add summary section at the bottom
    summary_records = []
    for sector, sgrp in sector_season.groupby("sector"):
        sorted_months = sgrp.sort_values("avg_return_pct", ascending=False)
        best_3 = ", ".join(sorted_months.head(3)["month_name"].tolist())
        worst_3 = ", ".join(sorted_months.tail(3)["month_name"].tolist())
        summary_records.append({
            "sector": sector,
            "month": "",
            "month_name": "SUMMARY",
            "avg_return_pct": None,
            "median_return_pct": None,
            "win_rate_pct": None,
            "sample_count": None,
            "best_stocks": f"Best: {best_3} | Worst: {worst_3}",
        })

    summary_df = pd.DataFrame(summary_records)
    sector_season = pd.concat([sector_season, summary_df], ignore_index=True)

    print(f"  Sector seasonality: {len(sector_records)} rows")
    print(f"  Stock seasonality: {len(stock_season)} rows")

    return sector_season, stock_season


def classify_trajectory(current, floor, ceiling, last_5_values):
    """Classify indicator trajectory relative to its historical floor."""
    if np.isnan(current) or np.isnan(floor) or np.isnan(ceiling):
        return "NEUTRAL"

    range_val = ceiling - floor
    if range_val <= 0:
        return "NEUTRAL"

    distance_pct = (current - floor) / range_val * 100

    # Check trend from last 5 values
    valid_vals = [v for v in last_5_values if not np.isnan(v)]
    if len(valid_vals) >= 3:
        trend = valid_vals[-1] - valid_vals[0]
        is_falling = trend < 0
        is_rising = trend > 0
    else:
        is_falling = False
        is_rising = False

    if distance_pct <= 10:
        return "AT_FLOOR"
    elif distance_pct <= 30 and is_falling:
        return "FALLING_TOWARD"
    elif distance_pct <= 15:
        # Check if was near floor in last 10 days — approximate with is_rising
        if is_rising:
            return "BOUNCING_UP"
        return "AT_FLOOR"
    elif distance_pct > 30 and is_rising:
        return "RISING_AWAY"
    else:
        return "NEUTRAL"


def build_indicator_floors(all_prices, sector_map, category_map):
    """Build Sheet 3: Indicator Floors (Capitulation) for A-category stocks."""
    print("[EXCEL] Computing indicator floors...")

    indicators = ["rsi", "macd_hist", "stoch_rsi", "mfi", "williams_r", "cmf", "bb_pct"]
    records = []

    a_stocks = {s for s, c in category_map.items() if c == "A"}
    processed = 0

    for sym, df in all_prices.items():
        if sym not in a_stocks or sym not in sector_map:
            continue

        df_ind = compute_all_indicators(df)

        if len(df_ind) < 50:
            continue

        processed += 1
        if processed % 20 == 0:
            print(f"  Processed {processed} stocks...")

        current_price = round(df_ind.iloc[-1]["close"], 1)
        row = {
            "symbol": sym,
            "sector": sector_map[sym],
            "current_price": current_price,
        }

        cap_score = 0
        bounce_potentials = []

        for ind in indicators:
            series = df_ind[ind].dropna()
            if len(series) < 20:
                for suffix in ["_floor", "_floor_date", "_current", "_distance_pct", "_trajectory", "_avg_bounce_pct"]:
                    row[f"{ind}{suffix}"] = None
                continue

            floor_val = series.min()
            ceiling_val = series.max()
            current_val = series.iloc[-1]
            floor_idx = series.idxmin()
            floor_date = df_ind.loc[floor_idx, "date"]

            range_val = ceiling_val - floor_val
            if range_val > 0:
                distance_pct = round((current_val - floor_val) / range_val * 100, 1)
            else:
                distance_pct = 50.0

            # Last 5 values for trend
            last_5 = series.tail(5).values.tolist()

            # Check if was near floor in last 10 days for BOUNCING_UP
            last_10 = series.tail(10)
            was_near_floor = False
            if range_val > 0:
                near_floor_mask = (last_10 - floor_val) / range_val * 100 <= 15
                was_near_floor = near_floor_mask.any()

            trajectory = classify_trajectory(current_val, floor_val, ceiling_val, last_5)
            # Override: if was near floor in last 10 and now rising, BOUNCING_UP
            if was_near_floor and trajectory == "NEUTRAL":
                valid_last5 = [v for v in last_5 if not np.isnan(v)]
                if len(valid_last5) >= 2 and valid_last5[-1] > valid_last5[0]:
                    trajectory = "BOUNCING_UP"

            # Avg bounce: forward 10-day return when indicator in bottom 10th percentile
            threshold = series.quantile(0.10)
            bottom_dates_idx = series[series <= threshold].index.tolist()
            forward_returns = []
            for idx in bottom_dates_idx:
                if idx + 10 < len(df_ind):
                    price_now = df_ind.loc[idx, "close"]
                    price_later = df_ind.loc[idx + 10, "close"]
                    if price_now > 0:
                        forward_returns.append((price_later / price_now - 1) * 100)
            avg_bounce = round(np.mean(forward_returns), 2) if forward_returns else 0.0

            row[f"{ind}_floor"] = round(floor_val, 4)
            row[f"{ind}_floor_date"] = str(floor_date)[:10]
            row[f"{ind}_current"] = round(current_val, 4)
            row[f"{ind}_distance_pct"] = distance_pct
            row[f"{ind}_trajectory"] = trajectory
            row[f"{ind}_avg_bounce_pct"] = avg_bounce

            if trajectory in ("AT_FLOOR", "FALLING_TOWARD"):
                cap_score += 1
            bounce_potentials.append(avg_bounce)

        row["capitulation_score"] = cap_score
        row["bounce_potential"] = round(np.mean(bounce_potentials), 2) if bounce_potentials else 0.0
        records.append(row)

    floors_df = pd.DataFrame(records)
    if not floors_df.empty:
        floors_df = floors_df.sort_values(
            ["capitulation_score", "bounce_potential"], ascending=[False, False]
        ).reset_index(drop=True)

    print(f"  Indicator floors computed for {len(records)} stocks")
    return floors_df


def build_capitulation_radar(floors_df):
    """Build Sheet 4: Top 30 by capitulation_score."""
    indicators = ["rsi", "macd_hist", "stoch_rsi", "mfi", "williams_r", "cmf", "bb_pct"]

    records = []
    for _, row in floors_df.head(30).iterrows():
        at_floor = []
        falling = []
        for ind in indicators:
            traj = row.get(f"{ind}_trajectory")
            if traj == "AT_FLOOR":
                at_floor.append(ind)
            elif traj == "FALLING_TOWARD":
                falling.append(ind)

        cap_score = row["capitulation_score"]
        if cap_score >= 5:
            overall = "CAPITULATING"
        elif cap_score >= 3:
            overall = "NEAR_BOTTOM"
        elif any(row.get(f"{ind}_trajectory") == "BOUNCING_UP" for ind in indicators):
            overall = "RECOVERING"
        else:
            overall = "NORMAL"

        records.append({
            "symbol": row["symbol"],
            "sector": row["sector"],
            "price": row["current_price"],
            "capitulation_score": cap_score,
            "indicators_at_floor": ", ".join(at_floor) if at_floor else "-",
            "indicators_falling": ", ".join(falling) if falling else "-",
            "overall_bounce_potential": row["bounce_potential"],
            "trajectory": overall,
        })

    return pd.DataFrame(records)


def build_sector_capitulation(floors_df):
    """Build Sheet 5: Sector-level capitulation summary."""
    if floors_df.empty:
        return pd.DataFrame()

    records = []
    for sector, grp in floors_df.groupby("sector"):
        top_cap = grp.nlargest(3, "capitulation_score")
        top_stocks = ", ".join(
            f"{r['symbol']}({r['capitulation_score']})" for _, r in top_cap.iterrows()
        )

        records.append({
            "sector": sector,
            "stock_count": len(grp),
            "avg_capitulation_score": round(grp["capitulation_score"].mean(), 2),
            "stocks_at_floor_count": (grp["capitulation_score"] >= 5).sum(),
            "stocks_falling_count": (
                (grp["capitulation_score"] >= 2) & (grp["capitulation_score"] < 5)
            ).sum(),
            "avg_bounce_potential": round(grp["bounce_potential"].mean(), 2),
            "top_capitulating_stocks": top_stocks,
        })

    return pd.DataFrame(records).sort_values("avg_capitulation_score", ascending=False)


def write_excel(sector_season, stock_season, floors_df, radar_df, sector_cap_df):
    """Write all sheets to Excel with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[EXCEL] Writing to {OUTPUT_PATH}...")

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        sector_season.to_excel(writer, sheet_name="Sector Seasonality", index=False)
        stock_season.to_excel(writer, sheet_name="Stock Seasonality", index=False)
        floors_df.to_excel(writer, sheet_name="Indicator Floors", index=False)
        radar_df.to_excel(writer, sheet_name="Capitulation Radar", index=False)
        sector_cap_df.to_excel(writer, sheet_name="Sector Capitulation", index=False)

        wb = writer.book

        bold_font = Font(bold=True)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        dark_green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Bold headers
            for cell in ws[1]:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="center")

            # Auto-width
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        max_len = max(max_len, len(val))
                    except Exception:
                        pass
                adjusted = min(max_len + 2, 40)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        # Color trajectory cells in Indicator Floors sheet
        if "Indicator Floors" in wb.sheetnames:
            ws = wb["Indicator Floors"]
            header_row = [cell.value for cell in ws[1]]
            traj_cols = [i for i, h in enumerate(header_row) if h and "_trajectory" in str(h)]

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in traj_cols:
                    cell = row[col_idx]
                    if cell.value == "AT_FLOOR":
                        cell.fill = green_fill
                    elif cell.value == "BOUNCING_UP":
                        cell.fill = green_fill
                    elif cell.value == "FALLING_TOWARD":
                        cell.fill = red_fill
                    elif cell.value == "NEUTRAL":
                        cell.fill = grey_fill

                # Color capitulation_score
                cap_col = next((i for i, h in enumerate(header_row) if h == "capitulation_score"), None)
                if cap_col is not None:
                    cell = row[cap_col]
                    try:
                        val = int(cell.value or 0)
                        if val >= 5:
                            cell.fill = dark_green_fill
                        elif val >= 3:
                            cell.fill = green_fill
                    except (ValueError, TypeError):
                        pass

        # Color Capitulation Radar
        if "Capitulation Radar" in wb.sheetnames:
            ws = wb["Capitulation Radar"]
            header_row = [cell.value for cell in ws[1]]
            traj_col = next((i for i, h in enumerate(header_row) if h == "trajectory"), None)

            if traj_col is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell = row[traj_col]
                    if cell.value == "CAPITULATING":
                        cell.fill = dark_green_fill
                    elif cell.value == "NEAR_BOTTOM":
                        cell.fill = green_fill
                    elif cell.value == "RECOVERING":
                        cell.fill = PatternFill(
                            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                        )

    print(f"[EXCEL] Written successfully!")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    start = time.time()

    # Phase 1: Backfill
    print("=" * 60)
    print("PHASE 1: Backfilling daily_prices from bdshare")
    print("=" * 60)
    backfill_prices()

    # Phase 2 & 3: Load data, compute indicators, build Excel
    print("\n" + "=" * 60)
    print("PHASE 2: Loading prices and computing indicators")
    print("=" * 60)
    all_prices = load_all_prices()
    sector_map, category_map = load_sector_map()

    print("\n" + "=" * 60)
    print("PHASE 3: Generating Excel")
    print("=" * 60)

    sector_season, stock_season = build_seasonality_sheets(all_prices, sector_map)
    floors_df = build_indicator_floors(all_prices, sector_map, category_map)
    radar_df = build_capitulation_radar(floors_df)
    sector_cap_df = build_sector_capitulation(floors_df)

    write_excel(sector_season, stock_season, floors_df, radar_df, sector_cap_df)

    elapsed = time.time() - start
    print(f"\n{'=' * 60}")
    print(f"DONE in {elapsed:.0f}s")
    print(f"Output: {OUTPUT_PATH}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
