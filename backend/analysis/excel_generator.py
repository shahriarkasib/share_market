"""Generate daily analysis Excel files.

Produces a multi-sheet workbook with BUY/WAIT/AVOID picks,
execution plans, technical data, and a trading checklist.
"""

import json
import logging
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Styles ───
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_BUY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_MACD_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
_DIP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_WAIT_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_AVOID_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_SECTION_FONT = Font(bold=True, size=13, color="2F5496")
_SUB_FONT = Font(bold=True, size=11, color="2F5496")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _fill_for_action(action: str) -> PatternFill | None:
    a = action.upper()
    if a == "BUY":
        return _BUY_FILL
    if "MACD" in a:
        return _MACD_FILL
    if "DIP" in a:
        return _DIP_FILL
    if "HOLD" in a or "WAIT" in a:
        return _WAIT_FILL
    if "AVOID" in a or "SELL" in a:
        return _AVOID_FILL
    return None


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if fill:
            cell.fill = fill


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_analysis_excel(data: list[dict], output) -> None:
    """Generate Excel workbook from daily analysis data.

    Args:
        data: list of analysis dicts (from daily_report or DB)
        output: file path string or BytesIO buffer
    """
    # Categorize
    buys = [s for s in data if s.get("action") == "BUY"]
    macd_wait = [s for s in data if "MACD" in str(s.get("action", ""))]
    dip_buy = [s for s in data if "dip" in str(s.get("action", "")).lower()]
    hold = [s for s in data if "HOLD" in str(s.get("action", "")) or "WAIT" in str(s.get("action", ""))]
    avoid = [s for s in data if "AVOID" in str(s.get("action", "")) or "SELL" in str(s.get("action", ""))]

    wb = openpyxl.Workbook()

    # ════════ Sheet 1: Summary Dashboard ════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_properties.tabColor = "2F5496"

    headers = ["#", "Symbol", "LTP", "Action", "Entry Range", "SL", "T1", "T2",
               "Risk%", "Reward%", "RSI", "StochRSI", "MACD", "Wait", "Vol Check", "Reasoning"]
    cols = len(headers)

    r = 1
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    ws1.cell(row=r, column=1, value=f"DSE Daily Analysis — {len(data)} Stocks").font = Font(bold=True, size=16, color="2F5496")
    r += 1
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    ws1.cell(row=r, column=1, value=f"BUY: {len(buys)} | MACD wait: {len(macd_wait)} | Buy on dip: {len(dip_buy)} | HOLD: {len(hold)} | AVOID: {len(avoid)}").font = Font(italic=True, size=10, color="666666")
    r += 2

    def _write_group(ws, r, title, stocks, fill):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        ws.cell(row=r, column=1, value=title).font = _SECTION_FONT
        r += 1
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=h)
        _style_header(ws, r, cols)
        r += 1
        for i, s in enumerate(stocks, 1):
            row_data = [
                i, s.get("symbol"), s.get("ltp"),
                s.get("action"),
                f"{s.get('entry_low', 0)}-{s.get('entry_high', 0)}",
                s.get("sl"), s.get("t1"), s.get("t2"),
                s.get("risk_pct"), s.get("reward_pct"),
                s.get("rsi"), s.get("stoch_rsi"),
                s.get("macd_status", ""),
                s.get("wait_days", ""),
                s.get("vol_entry", ""),
                s.get("reasoning", ""),
            ]
            for c, v in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=v)
            _style_row(ws, r, cols, fill)
            r += 1
        return r + 1

    r = _write_group(ws1, r, f"BUY NOW ({len(buys)})", buys, _BUY_FILL)
    r = _write_group(ws1, r, f"BUY — Wait for MACD ({len(macd_wait)})", macd_wait, _MACD_FILL)
    r = _write_group(ws1, r, f"BUY ON DIP ({len(dip_buy)})", dip_buy, _DIP_FILL)
    r = _write_group(ws1, r, f"HOLD/WAIT ({len(hold)})", hold, _WAIT_FILL)
    r = _write_group(ws1, r, f"SELL/AVOID ({len(avoid)})", avoid, _AVOID_FILL)

    _set_widths(ws1, [4, 14, 8, 22, 14, 8, 8, 8, 7, 8, 6, 8, 12, 12, 14, 55])

    # ════════ Sheet 2: Execution Plans ════════
    ws2 = wb.create_sheet("Execution Plans")
    ws2.sheet_properties.tabColor = "00B050"

    all_buys = buys + macd_wait + dip_buy
    r = 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws2.cell(row=r, column=1, value=f"EXECUTION PLANS — {len(all_buys)} Actionable Stocks").font = Font(bold=True, size=16, color="00B050")
    r += 2

    for s in all_buys:
        # Header
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws2.cell(row=r, column=1, value=f"{s.get('symbol')} — {s.get('action')} (LTP: {s.get('ltp')})")
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        r += 1

        # Metrics
        metrics = [
            ["Entry Range", f"{s.get('entry_low')}-{s.get('entry_high')}", "RSI", s.get("rsi"), "MACD Status", s.get("macd_status"), "Wait", s.get("wait_days")],
            ["Stop Loss", f"{s.get('sl')} ({s.get('risk_pct')}% risk)", "StochRSI", s.get("stoch_rsi"), "MACD Hist", s.get("macd_hist"), "ATR", f"{s.get('atr')} ({s.get('atr_pct')}%)"],
            ["Target 1", s.get("t1"), "BB Position", f"{round(s.get('bb_pct', 0) * 100, 1)}%", "Trend 50d", f"{s.get('trend_50d')}%", "Max DD", f"{s.get('max_dd')}%"],
            ["Target 2", f"{s.get('t2')} ({s.get('reward_pct')}% reward)", "Support", s.get("support"), "Resistance", s.get("resistance"), "Vol Entry", s.get("vol_entry")],
        ]
        for row_data in metrics:
            for c, v in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                if c % 2 == 1:
                    cell.font = Font(bold=True, size=9, color="666666")
                cell.border = _THIN_BORDER
            r += 1
        r += 1

        # Scenarios
        scenarios = s.get("scenarios_json")
        if isinstance(scenarios, str):
            try:
                scenarios = json.loads(scenarios)
            except (json.JSONDecodeError, TypeError):
                scenarios = []
        if scenarios:
            ws2.cell(row=r, column=1, value="ENTRY SCENARIOS").font = _SUB_FONT
            r += 1
            for sc in scenarios:
                ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                ws2.cell(row=r, column=1, value=sc.get("name", "")).font = Font(bold=True, size=10, color="C65911")
                r += 1
                for step in sc.get("steps", []):
                    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                    ws2.cell(row=r, column=1, value=f"  {step}").alignment = _WRAP
                    r += 1
            r += 1

        # Reasoning summary
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws2.cell(row=r, column=1, value=s.get("reasoning", ""))
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.alignment = _WRAP
        r += 3

    _set_widths(ws2, [18, 16, 18, 16, 18, 16, 18, 16])

    # ════════ Sheet 3: HOLD/WAIT Details ════════
    ws3 = wb.create_sheet("HOLD-WAIT")
    ws3.sheet_properties.tabColor = "FFC000"
    r = 1
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws3.cell(row=r, column=1, value=f"HOLD/WAIT ({len(hold)}) — When They Become Buyable").font = Font(bold=True, size=16, color="FFC000")
    r += 2
    h3 = ["#", "Symbol", "LTP", "Entry (when ready)", "SL", "T1", "T2", "Wait", "Reasoning"]
    for c, h in enumerate(h3, 1):
        ws3.cell(row=r, column=c, value=h)
    _style_header(ws3, r, len(h3))
    r += 1
    for i, s in enumerate(hold, 1):
        for c, v in enumerate([i, s.get("symbol"), s.get("ltp"),
                               f"{s.get('entry_low')}-{s.get('entry_high')}", s.get("sl"),
                               s.get("t1"), s.get("t2"), s.get("wait_days"), s.get("reasoning")], 1):
            ws3.cell(row=r, column=c, value=v)
        _style_row(ws3, r, len(h3), _WAIT_FILL)
        r += 1
    _set_widths(ws3, [4, 14, 8, 16, 8, 8, 8, 12, 60])

    # ════════ Sheet 4: AVOID Details ════════
    ws4 = wb.create_sheet("AVOID")
    ws4.sheet_properties.tabColor = "FF0000"
    r = 1
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws4.cell(row=r, column=1, value=f"SELL/AVOID ({len(avoid)})").font = Font(bold=True, size=16, color="FF0000")
    r += 2
    h4 = ["#", "Symbol", "LTP", "Entry IF corrects", "SL", "T1", "T2", "Wait", "Reasoning"]
    for c, h in enumerate(h4, 1):
        ws4.cell(row=r, column=c, value=h)
    _style_header(ws4, r, len(h4))
    r += 1
    for i, s in enumerate(avoid, 1):
        for c, v in enumerate([i, s.get("symbol"), s.get("ltp"),
                               f"{s.get('entry_low')}-{s.get('entry_high')}", s.get("sl"),
                               s.get("t1"), s.get("t2"), s.get("wait_days"), s.get("reasoning")], 1):
            ws4.cell(row=r, column=c, value=v)
        _style_row(ws4, r, len(h4), _AVOID_FILL)
        r += 1
    _set_widths(ws4, [4, 14, 8, 16, 8, 8, 8, 12, 60])

    # ════════ Sheet 5: All Technicals ════════
    ws5 = wb.create_sheet("Technicals")
    ws5.sheet_properties.tabColor = "7030A0"
    tech_h = ["Symbol", "Action", "LTP", "RSI", "StochRSI", "MACD Hist", "BB%",
              "EMA9", "EMA21", "SMA50", "ATR%", "Vol%", "MaxDD", "5d%", "10d%",
              "Trend50d", "Support", "Resist", "WinRate", "Bounce", "AvgVol", "VolRatio"]
    r = 1
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(tech_h))
    ws5.cell(row=r, column=1, value="FULL TECHNICAL DATA").font = Font(bold=True, size=16, color="7030A0")
    r += 2
    for c, h in enumerate(tech_h, 1):
        ws5.cell(row=r, column=c, value=h)
    _style_header(ws5, r, len(tech_h))
    r += 1
    for s in sorted(data, key=lambda x: x.get("rsi", 50)):
        row_data = [
            s.get("symbol"), s.get("action"), s.get("ltp"),
            s.get("rsi"), s.get("stoch_rsi"), s.get("macd_hist"),
            f"{round(s.get('bb_pct', 0) * 100, 1)}%",
            s.get("ema9"), s.get("ema21"), s.get("sma50"),
            f"{s.get('atr_pct')}%", f"{s.get('volatility')}%", f"{s.get('max_dd')}%",
            f"{s.get('chg_5d', 0)}%", f"{s.get('chg_10d', 0)}%", f"{s.get('trend_50d')}%",
            s.get("support"), s.get("resistance"),
            f"{s.get('win_rate', 0)}%", f"{s.get('bounce_rate', 0)}%",
            s.get("avg_vol"), s.get("vol_ratio"),
        ]
        for c, v in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=v)
        _style_row(ws5, r, len(tech_h), _fill_for_action(s.get("action", "")))
        r += 1
    _set_widths(ws5, [14, 22, 8, 6, 8, 8, 7, 8, 8, 8, 7, 7, 7, 6, 6, 8, 8, 8, 7, 7, 10, 7])

    # Save
    wb.save(output)
    logger.info(f"Generated analysis Excel with {len(data)} stocks, 5 sheets")
