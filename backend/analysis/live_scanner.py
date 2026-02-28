"""Intraday live scanner — fetches market depth for BUY-signal stocks,
computes buy pressure, assesses T+2 risk, logs decisions for backtesting,
and appends results to a running Excel file.

Runs every 5 minutes during market hours (9:55-14:30 Sun-Thu).
"""

import logging
import math
import os
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import pytz
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import get_connection

logger = logging.getLogger(__name__)
DSE_TZ = pytz.timezone("Asia/Dhaka")

# ─── Excel styles ───
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_BUY_NOW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_READY_FILL = PatternFill(start_color="B4E6C8", end_color="B4E6C8", fill_type="solid")
_ACCUMULATE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_BOOK_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
_EXIT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_WAIT_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_WATCH_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_REC_FILL = {
    "BUY NOW": _BUY_NOW_FILL,
    "READY": _READY_FILL,
    "ACCUMULATE": _ACCUMULATE_FILL,
    "WEAK-WAIT": _WAIT_FILL,
    "BOOK PARTIAL": _BOOK_FILL,
    "BOOK FULL": _BOOK_FILL,
    "EXIT": _EXIT_FILL,
    "WATCH": _WATCH_FILL,
}

# Excel output directory
SCAN_DIR = os.environ.get("LIVE_SCAN_DIR", "/tmp")

# In-memory cache of latest scan results (for API)
_latest_scan: dict = {"timestamp": None, "results": [], "summary": {}}


def _compute_status(ltp: float, entry_low: float, entry_high: float,
                    sl: float, t1: float, t2: float) -> tuple[str, float]:
    """Compute tracking status and distance % from entry zone midpoint."""
    entry_mid = (entry_low + entry_high) / 2 if entry_low > 0 else ltp
    dist_pct = round((ltp - entry_mid) / entry_mid * 100, 1) if entry_mid > 0 else 0

    if sl > 0 and ltp <= sl:
        return "SL_HIT", dist_pct
    if t2 > 0 and ltp >= t2:
        return "T2_HIT", dist_pct
    if t1 > 0 and ltp >= t1:
        return "T1_HIT", dist_pct
    if entry_low > 0 and entry_high > 0 and entry_low <= ltp <= entry_high:
        return "ENTRY_ZONE", dist_pct
    if entry_low > 0 and ltp < entry_low and (sl <= 0 or ltp > sl):
        return "BELOW_ENTRY", dist_pct
    if entry_high > 0 and ltp <= entry_high * 1.02:
        return "APPROACHING", dist_pct
    return "WATCHING", dist_pct


def _compute_t2_risk(*, rsi: float, stoch_rsi: float, macd_status: str,
                     macd_hist: float, trend_50d: float, volatility: float,
                     atr_pct: float, risk_pct: float, vol_ratio: float,
                     live_change_pct: float) -> tuple[str, str]:
    """Assess T+2 holding risk — will the stock hold value for 2+ trading days?

    Returns (risk_level, reason).
    risk_level: "LOW", "MEDIUM", "HIGH"
    """
    risk_score = 0
    reasons = []

    # 1. MACD direction — bearish MACD means momentum is fading
    macd_lower = (macd_status or "").lower()
    if "bearish" in macd_lower:
        risk_score += 25
        reasons.append("MACD bearish")
    elif "cross" in macd_lower and "down" in macd_lower:
        risk_score += 20
        reasons.append("MACD crossing down")

    # Negative MACD histogram = selling pressure building
    if macd_hist < -0.5:
        risk_score += 10
        reasons.append(f"MACD hist negative ({macd_hist:.1f})")

    # 2. RSI — overbought (>70) means likely pullback within T+2
    if rsi > 75:
        risk_score += 25
        reasons.append(f"RSI overbought ({rsi:.0f})")
    elif rsi > 65:
        risk_score += 10
        reasons.append(f"RSI elevated ({rsi:.0f})")
    # Very oversold could bounce — actually lower risk for T+2
    elif rsi < 25:
        risk_score -= 10

    # 3. StochRSI — overbought stoch = reversal likely
    if stoch_rsi > 85:
        risk_score += 15
        reasons.append(f"StochRSI overbought ({stoch_rsi:.0f})")

    # 4. Trend — downtrend means stock is already falling
    if trend_50d < -5:
        risk_score += 20
        reasons.append(f"50d downtrend ({trend_50d:.1f}%)")
    elif trend_50d < -2:
        risk_score += 10
        reasons.append(f"Weak trend ({trend_50d:.1f}%)")
    elif trend_50d > 5:
        risk_score -= 10  # Strong uptrend = lower T+2 risk

    # 5. Volatility — high volatility = bigger swings, might hit SL before T+2
    if volatility > 4:
        risk_score += 15
        reasons.append(f"High volatility ({volatility:.1f}%)")
    elif atr_pct > 3:
        risk_score += 10
        reasons.append(f"High ATR ({atr_pct:.1f}%)")

    # 6. Risk % — if SL is far from entry, more room to drop
    if risk_pct > 5:
        risk_score += 10
        reasons.append(f"Wide SL ({risk_pct:.1f}%)")

    # 7. Today's change — if already up big today, might mean-revert by T+2
    if live_change_pct > 5:
        risk_score += 15
        reasons.append(f"Already up {live_change_pct:.1f}% today")
    elif live_change_pct < -3:
        risk_score += 10
        reasons.append(f"Falling {live_change_pct:.1f}% today")

    # 8. Low volume = less conviction
    if vol_ratio < 0.5:
        risk_score += 10
        reasons.append(f"Low volume ({vol_ratio:.1f}x avg)")

    # Classify
    if risk_score >= 40:
        return "HIGH", "; ".join(reasons) or "Multiple risk factors"
    elif risk_score >= 20:
        return "MEDIUM", "; ".join(reasons) or "Some risk factors"
    else:
        return "LOW", "; ".join(reasons) if reasons else "Favorable conditions for T+2 hold"


def _fetch_market_depth(symbol: str) -> dict:
    """Fetch order book (buy/sell depth) for a symbol from DSE.

    Returns dict with keys: buy_levels, sell_levels, total_buy_vol,
    total_sell_vol, best_bid, best_ask, buy_sell_ratio, spread_pct.
    """
    try:
        from bdshare import get_market_depth_data
        df = get_market_depth_data(symbol)

        if df is None or df.empty:
            return {"empty": True, "total_buy_vol": 0, "total_sell_vol": 0,
                    "buy_sell_ratio": 0, "best_bid": 0, "best_ask": 0,
                    "spread_pct": 0, "buy_levels": 0, "sell_levels": 0}

        total_buy_vol = 0
        total_sell_vol = 0
        best_bid = 0
        best_ask = float("inf")
        buy_levels = 0
        sell_levels = 0

        for _, row in df.iterrows():
            bv = float(row.get("buy_volume", 0) or 0)
            bp = float(row.get("buy_price", 0) or 0)
            sp = float(row.get("sell_price", 0) or 0)
            sv = float(row.get("sell_volume", 0) or 0)

            if bv > 0:
                total_buy_vol += bv
                buy_levels += 1
                if bp > best_bid:
                    best_bid = bp

            if sv > 0:
                total_sell_vol += sv
                sell_levels += 1
                if sp > 0 and sp < best_ask:
                    best_ask = sp

        if best_ask == float("inf"):
            best_ask = 0

        ratio = round(total_buy_vol / total_sell_vol, 2) if total_sell_vol > 0 else (
            99.0 if total_buy_vol > 0 else 0
        )
        spread_pct = round((best_ask - best_bid) / best_bid * 100, 2) if best_bid > 0 and best_ask > 0 else 0

        return {
            "empty": False,
            "total_buy_vol": int(total_buy_vol),
            "total_sell_vol": int(total_sell_vol),
            "buy_sell_ratio": ratio,
            "best_bid": round(best_bid, 1),
            "best_ask": round(best_ask, 1),
            "spread_pct": spread_pct,
            "buy_levels": buy_levels,
            "sell_levels": sell_levels,
        }
    except Exception as e:
        logger.warning(f"Depth fetch failed for {symbol}: {e}")
        return {"empty": True, "total_buy_vol": 0, "total_sell_vol": 0,
                "buy_sell_ratio": 0, "best_bid": 0, "best_ask": 0,
                "spread_pct": 0, "buy_levels": 0, "sell_levels": 0}


def _decide_recommendation(status: str, buy_sell_ratio: float,
                           spread_pct: float, depth_empty: bool,
                           t2_risk: str) -> tuple[str, str]:
    """Decide recommendation based on status + order book + T+2 risk.

    Returns (recommendation, reason).
    """
    if status == "SL_HIT":
        return "EXIT", "Price hit stop loss"
    if status == "T2_HIT":
        return "BOOK FULL", "Price reached Target 2"
    if status == "T1_HIT":
        return "BOOK PARTIAL", "Price reached Target 1 — book partial profits"

    if depth_empty:
        reason_suffix = " (no depth data)"
    else:
        reason_suffix = ""

    # T+2 risk override: downgrade BUY NOW to WEAK-WAIT if T+2 risk is HIGH
    if status == "ENTRY_ZONE":
        if t2_risk == "HIGH":
            return "WEAK-WAIT", f"In entry zone but HIGH T+2 risk — may drop before you can sell{reason_suffix}"
        if buy_sell_ratio > 1.5:
            return "BUY NOW", f"In entry zone, strong buy pressure (ratio {buy_sell_ratio}x), T+2 risk {t2_risk}{reason_suffix}"
        if buy_sell_ratio > 1.0:
            return "BUY NOW", f"In entry zone, moderate buy pressure (ratio {buy_sell_ratio}x), T+2 risk {t2_risk}{reason_suffix}"
        if buy_sell_ratio > 0.8:
            return "WEAK-WAIT", f"In entry zone but sellers dominating (ratio {buy_sell_ratio}x){reason_suffix}"
        return "WEAK-WAIT", f"In entry zone but heavy sell pressure (ratio {buy_sell_ratio}x){reason_suffix}"

    if status == "APPROACHING":
        if t2_risk == "HIGH":
            return "WATCH", f"Near entry but HIGH T+2 risk — wait for momentum shift{reason_suffix}"
        if buy_sell_ratio > 2.0:
            return "READY", f"Near entry + very strong buy pressure (ratio {buy_sell_ratio}x) — place limit order{reason_suffix}"
        if buy_sell_ratio > 1.2:
            return "READY", f"Near entry + buyers leading (ratio {buy_sell_ratio}x){reason_suffix}"
        return "WATCH", f"Approaching entry but buy pressure weak (ratio {buy_sell_ratio}x){reason_suffix}"

    if status == "BELOW_ENTRY":
        if buy_sell_ratio > 1.0:
            return "ACCUMULATE", f"Below entry zone, buyers present (ratio {buy_sell_ratio}x) — potential bounce{reason_suffix}"
        return "WATCH", f"Below entry, sellers dominating (ratio {buy_sell_ratio}x){reason_suffix}"

    return "WATCH", f"Above entry zone, waiting{reason_suffix}"


def _safe_float(v) -> float:
    if v is None:
        return 0
    f = float(v)
    return 0 if (math.isnan(f) or math.isinf(f)) else f


def scan_buy_signals() -> list[dict]:
    """Load today's BUY-type analysis, fetch depth, compute recommendations."""
    now = datetime.now(DSE_TZ)
    conn = get_connection()

    # Get latest analysis date
    row = conn.execute("SELECT MAX(date) FROM daily_analysis").fetchone()
    analysis_date = str(row[0]) if row and row[0] else now.strftime("%Y-%m-%d")

    # Load BUY-type stocks with live prices + extra fields for T+2 risk
    rows = conn.execute(
        """SELECT da.symbol, da.action, da.entry_low, da.entry_high,
                  da.sl, da.t1, da.t2, da.score, da.category,
                  da.rsi, da.stoch_rsi, da.macd_status, da.macd_hist,
                  da.risk_pct, da.reward_pct,
                  da.trend_50d, da.volatility, da.atr_pct, da.vol_ratio,
                  lp.ltp AS live_ltp, lp.change_pct AS live_change_pct,
                  lp.volume AS live_volume, lp.high AS live_high, lp.low AS live_low,
                  f.sector
           FROM daily_analysis da
           JOIN live_prices lp ON da.symbol = lp.symbol
           LEFT JOIN fundamentals f ON da.symbol = f.symbol
           WHERE da.date = %s
             AND da.action LIKE 'BUY%%'
           ORDER BY da.score DESC
        """,
        (analysis_date,),
    )
    buy_stocks = rows.fetchall()
    conn.close()

    results = []
    scan_time = now.strftime("%H:%M:%S")

    for r in buy_stocks:
        symbol = r["symbol"]
        ltp = float(r["live_ltp"] or 0)
        if ltp <= 0:
            continue

        entry_low = float(r["entry_low"] or 0)
        entry_high = float(r["entry_high"] or 0)
        sl = float(r["sl"] or 0)
        t1 = float(r["t1"] or 0)
        t2 = float(r["t2"] or 0)

        # Compute tracking status
        status, dist_pct = _compute_status(ltp, entry_low, entry_high, sl, t1, t2)

        # Compute T+2 risk
        t2_risk, t2_reason = _compute_t2_risk(
            rsi=_safe_float(r["rsi"]),
            stoch_rsi=_safe_float(r["stoch_rsi"]),
            macd_status=r["macd_status"] or "",
            macd_hist=_safe_float(r.get("macd_hist")),
            trend_50d=_safe_float(r.get("trend_50d")),
            volatility=_safe_float(r.get("volatility")),
            atr_pct=_safe_float(r.get("atr_pct")),
            risk_pct=_safe_float(r["risk_pct"]),
            vol_ratio=_safe_float(r.get("vol_ratio")),
            live_change_pct=_safe_float(r["live_change_pct"]),
        )

        # Fetch market depth with rate limiting
        depth = _fetch_market_depth(symbol)
        time.sleep(0.3)  # Rate limit: ~3 req/sec

        # Decide recommendation (now considers T+2 risk)
        rec, reason = _decide_recommendation(
            status, depth["buy_sell_ratio"], depth["spread_pct"],
            depth["empty"], t2_risk,
        )

        results.append({
            "timestamp": scan_time,
            "symbol": symbol,
            "action": r["action"],
            "category": r["category"] or "",
            "sector": r["sector"] or "",
            "score": _safe_float(r["score"]),
            "live_ltp": round(ltp, 1),
            "live_change_pct": _safe_float(r["live_change_pct"]),
            "live_volume": int(r["live_volume"] or 0),
            "entry_low": entry_low,
            "entry_high": entry_high,
            "sl": sl,
            "t1": t1,
            "t2": t2,
            "status": status,
            "distance_pct": dist_pct,
            "total_buy_vol": depth["total_buy_vol"],
            "total_sell_vol": depth["total_sell_vol"],
            "buy_sell_ratio": depth["buy_sell_ratio"],
            "best_bid": depth["best_bid"],
            "best_ask": depth["best_ask"],
            "spread_pct": depth["spread_pct"],
            "buy_levels": depth["buy_levels"],
            "sell_levels": depth["sell_levels"],
            "recommendation": rec,
            "reasoning": reason,
            "rsi": _safe_float(r["rsi"]),
            "macd_status": r["macd_status"] or "",
            "t2_risk": t2_risk,
            "t2_risk_reason": t2_reason,
        })

    # Sort: BUY NOW first, then READY, ACCUMULATE, etc.
    rec_priority = {
        "BUY NOW": 0, "READY": 1, "ACCUMULATE": 2, "WEAK-WAIT": 3,
        "BOOK PARTIAL": 4, "BOOK FULL": 5, "EXIT": 6, "WATCH": 7,
    }
    results.sort(key=lambda s: (rec_priority.get(s["recommendation"], 9), -s["score"]))

    return results


def _log_decisions(results: list[dict], scan_time_iso: str):
    """Save every scan decision to the database for backtesting."""
    if not results:
        return
    conn = get_connection()
    date_str = datetime.now(DSE_TZ).strftime("%Y-%m-%d")

    for r in results:
        try:
            conn.execute(
                """INSERT INTO scan_decisions
                   (date, scan_time, symbol, recommendation, live_ltp,
                    entry_low, entry_high, sl, t1, t2,
                    status, buy_sell_ratio, t2_risk, score,
                    rsi, macd_status, reasoning)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                           %s, %s, %s, %s, %s, %s, %s)
                   ON CONFLICT (date, scan_time, symbol) DO UPDATE SET
                     recommendation = EXCLUDED.recommendation,
                     live_ltp = EXCLUDED.live_ltp,
                     buy_sell_ratio = EXCLUDED.buy_sell_ratio,
                     t2_risk = EXCLUDED.t2_risk,
                     reasoning = EXCLUDED.reasoning
                """,
                (
                    date_str, scan_time_iso, r["symbol"], r["recommendation"],
                    r["live_ltp"], r["entry_low"], r["entry_high"],
                    r["sl"], r["t1"], r["t2"],
                    r["status"], r["buy_sell_ratio"], r["t2_risk"],
                    r["score"], r["rsi"], r["macd_status"], r["reasoning"],
                ),
            )
        except Exception as e:
            logger.error(f"Decision log error for {r['symbol']}: {e}")

    conn.commit()
    conn.close()
    logger.info(f"Logged {len(results)} scan decisions to database")


def verify_past_decisions():
    """Check past scan decisions against actual prices at T+1, T+2, T+3, T+5, T+7.

    Run daily after market close to fill in actual outcomes.
    """
    conn = get_connection()

    # Find decisions that need verification (no actual_t2 yet, at least 2 trading days old)
    rows = conn.execute(
        """SELECT DISTINCT sd.date, sd.symbol, sd.recommendation, sd.live_ltp,
                  sd.sl, sd.t1, sd.t2
           FROM scan_decisions sd
           WHERE sd.actual_t2 IS NULL
             AND sd.recommendation IN ('BUY NOW', 'READY', 'ACCUMULATE')
             AND sd.date < CURRENT_DATE - INTERVAL '1 day'
           ORDER BY sd.date
           LIMIT 500
        """
    ).fetchall()

    if not rows:
        logger.info("No decisions to verify")
        conn.close()
        return

    # Group by symbol to batch lookups
    symbols = set(r["symbol"] for r in rows)
    dates_needed = set(r["date"] for r in rows)

    logger.info(f"Verifying {len(rows)} decisions across {len(symbols)} symbols")

    for r in rows:
        symbol = r["symbol"]
        decision_date = str(r["date"])
        decision_ltp = float(r["live_ltp"])

        # Get closing prices for T+1 through T+7 trading days after decision
        price_rows = conn.execute(
            """SELECT date, close FROM daily_prices
               WHERE symbol = %s AND date > %s
               ORDER BY date ASC LIMIT 7
            """,
            (symbol, decision_date),
        ).fetchall()

        if not price_rows:
            continue

        prices = {i + 1: float(pr["close"]) for i, pr in enumerate(price_rows)}
        sl = float(r["sl"])
        t1 = float(r["t1"])
        t2_target = float(r["t2"])

        # Compute returns
        actual_t1 = prices.get(1)
        actual_t2 = prices.get(2)
        actual_t3 = prices.get(3)
        actual_t5 = prices.get(5)
        actual_t7 = prices.get(7)

        return_t2 = round((actual_t2 - decision_ltp) / decision_ltp * 100, 2) if actual_t2 else None

        # Did it hit SL or target within 7 days?
        sl_hit_day = None
        t1_hit_day = None
        t2_hit_day = None
        for day_n, price in sorted(prices.items()):
            if sl > 0 and price <= sl and sl_hit_day is None:
                sl_hit_day = day_n
            if t1 > 0 and price >= t1 and t1_hit_day is None:
                t1_hit_day = day_n
            if t2_target > 0 and price >= t2_target and t2_hit_day is None:
                t2_hit_day = day_n

        # Was the decision correct? (profitable at T+2)
        outcome = "UNKNOWN"
        if return_t2 is not None:
            if return_t2 > 0:
                outcome = "CORRECT"
            elif return_t2 > -1:
                outcome = "MARGINAL"
            else:
                outcome = "WRONG"

        conn.execute(
            """UPDATE scan_decisions SET
                 actual_t1 = %s, actual_t2 = %s, actual_t3 = %s,
                 actual_t5 = %s, actual_t7 = %s,
                 return_t2_pct = %s, outcome = %s,
                 sl_hit_day = %s, t1_hit_day = %s, t2_hit_day = %s
               WHERE date = %s AND symbol = %s
                 AND recommendation = %s AND actual_t2 IS NULL
            """,
            (
                actual_t1, actual_t2, actual_t3, actual_t5, actual_t7,
                return_t2, outcome,
                sl_hit_day, t1_hit_day, t2_hit_day,
                decision_date, symbol, r["recommendation"],
            ),
        )

    conn.commit()
    conn.close()
    logger.info(f"Verified {len(rows)} past decisions")


def get_decision_accuracy(days: int = 30) -> dict:
    """Get accuracy stats for past scan decisions.

    Returns summary for the LLM to learn from.
    """
    conn = get_connection()

    # Overall accuracy
    stats = conn.execute(
        """SELECT
             COUNT(*) as total,
             COUNT(CASE WHEN outcome = 'CORRECT' THEN 1 END) as correct,
             COUNT(CASE WHEN outcome = 'WRONG' THEN 1 END) as wrong,
             COUNT(CASE WHEN outcome = 'MARGINAL' THEN 1 END) as marginal,
             AVG(return_t2_pct) as avg_return_t2,
             COUNT(CASE WHEN sl_hit_day IS NOT NULL THEN 1 END) as sl_hits,
             COUNT(CASE WHEN t1_hit_day IS NOT NULL THEN 1 END) as t1_hits
           FROM scan_decisions
           WHERE outcome IS NOT NULL
             AND date >= CURRENT_DATE - INTERVAL '%s days'
             AND recommendation IN ('BUY NOW', 'READY', 'ACCUMULATE')
        """ % days,
    ).fetchone()

    # Per-recommendation accuracy
    by_rec = conn.execute(
        """SELECT recommendation,
             COUNT(*) as total,
             COUNT(CASE WHEN outcome = 'CORRECT' THEN 1 END) as correct,
             AVG(return_t2_pct) as avg_return,
             COUNT(CASE WHEN sl_hit_day <= 3 THEN 1 END) as early_sl_hits
           FROM scan_decisions
           WHERE outcome IS NOT NULL
             AND date >= CURRENT_DATE - INTERVAL '%s days'
           GROUP BY recommendation ORDER BY total DESC
        """ % days,
    ).fetchall()

    # Worst calls (biggest losses at T+2)
    worst = conn.execute(
        """SELECT symbol, date, recommendation, live_ltp, actual_t2,
                  return_t2_pct, t2_risk, reasoning
           FROM scan_decisions
           WHERE outcome = 'WRONG' AND return_t2_pct IS NOT NULL
             AND date >= CURRENT_DATE - INTERVAL '%s days'
           ORDER BY return_t2_pct ASC LIMIT 10
        """ % days,
    ).fetchall()

    # Best calls
    best = conn.execute(
        """SELECT symbol, date, recommendation, live_ltp, actual_t2,
                  return_t2_pct, t2_risk
           FROM scan_decisions
           WHERE outcome = 'CORRECT' AND return_t2_pct IS NOT NULL
             AND date >= CURRENT_DATE - INTERVAL '%s days'
           ORDER BY return_t2_pct DESC LIMIT 10
        """ % days,
    ).fetchall()

    conn.close()

    total = stats["total"] if stats else 0
    correct = stats["correct"] if stats else 0

    return {
        "total_verified": total,
        "accuracy_pct": round(correct / total * 100, 1) if total > 0 else 0,
        "correct": correct,
        "wrong": stats["wrong"] if stats else 0,
        "marginal": stats["marginal"] if stats else 0,
        "avg_return_t2": round(float(stats["avg_return_t2"] or 0), 2),
        "sl_hits": stats["sl_hits"] if stats else 0,
        "t1_hits": stats["t1_hits"] if stats else 0,
        "by_recommendation": [
            {
                "recommendation": r["recommendation"],
                "total": r["total"],
                "correct": r["correct"],
                "accuracy_pct": round(r["correct"] / r["total"] * 100, 1) if r["total"] > 0 else 0,
                "avg_return": round(float(r["avg_return"] or 0), 2),
                "early_sl_hits": r["early_sl_hits"],
            }
            for r in by_rec
        ],
        "worst_calls": [
            {
                "symbol": r["symbol"], "date": str(r["date"]),
                "recommendation": r["recommendation"],
                "ltp": float(r["live_ltp"]), "actual_t2": float(r["actual_t2"] or 0),
                "return_pct": float(r["return_t2_pct"] or 0),
                "t2_risk": r["t2_risk"], "reasoning": r["reasoning"],
            }
            for r in worst
        ],
        "best_calls": [
            {
                "symbol": r["symbol"], "date": str(r["date"]),
                "recommendation": r["recommendation"],
                "ltp": float(r["live_ltp"]), "actual_t2": float(r["actual_t2"] or 0),
                "return_pct": float(r["return_t2_pct"] or 0),
            }
            for r in best
        ],
    }


def _get_excel_path(date_str: str | None = None) -> str:
    """Get path to today's live scan Excel file."""
    if not date_str:
        date_str = datetime.now(DSE_TZ).strftime("%Y-%m-%d")
    return os.path.join(SCAN_DIR, f"DSE_LiveScan_{date_str}.xlsx")


def append_to_excel(scan_results: list[dict], filepath: str | None = None) -> str:
    """Append scan results to the running Excel file.

    Creates the file if it doesn't exist; appends rows if it does.
    Returns the filepath.
    """
    if not filepath:
        filepath = _get_excel_path()

    headers = [
        "Time", "Symbol", "Action", "LTP", "Chg%", "Entry Low", "Entry High",
        "SL", "T1", "T2", "Status", "Dist%", "Buy Vol", "Sell Vol",
        "B/S Ratio", "Best Bid", "Best Ask", "Spread%",
        "T+2 Risk", "Recommendation", "RSI", "MACD", "Score", "Reasoning",
    ]

    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Live Scan"] if "Live Scan" in wb.sheetnames else wb.active
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Live Scan"
        ws.sheet_properties.tabColor = "2F5496"
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER
        widths = [8, 12, 18, 8, 7, 8, 8, 8, 8, 8, 12, 7, 10, 10, 8, 8, 8, 7, 10, 16, 6, 10, 6, 45]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        next_row = 2

    for r in scan_results:
        row_data = [
            r["timestamp"], r["symbol"], r["action"], r["live_ltp"],
            r["live_change_pct"], r["entry_low"], r["entry_high"],
            r["sl"], r["t1"], r["t2"], r["status"], r["distance_pct"],
            r["total_buy_vol"], r["total_sell_vol"], r["buy_sell_ratio"],
            r["best_bid"], r["best_ask"], r["spread_pct"],
            r["t2_risk"], r["recommendation"], r["rsi"], r["macd_status"],
            r["score"], r["reasoning"],
        ]
        for c, v in enumerate(row_data, 1):
            ws.cell(row=next_row, column=c, value=v)

        fill = _REC_FILL.get(r["recommendation"], None)
        if fill:
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=next_row, column=c)
                cell.fill = fill
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(vertical="top")

        next_row += 1

    # Summary sheet
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.sheet_properties.tabColor = "00B050"

    scan_time = scan_results[0]["timestamp"] if scan_results else "N/A"
    ws_sum.merge_cells("A1:E1")
    ws_sum.cell(row=1, column=1, value=f"Live Scan Summary — {scan_time}").font = Font(
        bold=True, size=14, color="2F5496"
    )

    counts: dict[str, int] = {}
    for r in scan_results:
        rec = r["recommendation"]
        counts[rec] = counts.get(rec, 0) + 1

    row = 3
    ws_sum.cell(row=row, column=1, value="Recommendation").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value="Count").font = Font(bold=True)
    _style_header_cell(ws_sum.cell(row=row, column=1))
    _style_header_cell(ws_sum.cell(row=row, column=2))
    row += 1
    for rec_name in ["BUY NOW", "READY", "ACCUMULATE", "WEAK-WAIT",
                     "BOOK PARTIAL", "BOOK FULL", "EXIT", "WATCH"]:
        cnt = counts.get(rec_name, 0)
        if cnt > 0:
            ws_sum.cell(row=row, column=1, value=rec_name)
            ws_sum.cell(row=row, column=2, value=cnt)
            fill = _REC_FILL.get(rec_name)
            if fill:
                ws_sum.cell(row=row, column=1).fill = fill
                ws_sum.cell(row=row, column=2).fill = fill
            ws_sum.cell(row=row, column=1).border = _THIN_BORDER
            ws_sum.cell(row=row, column=2).border = _THIN_BORDER
            row += 1

    # T+2 risk breakdown
    row += 1
    ws_sum.cell(row=row, column=1, value="T+2 Risk").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value="Count").font = Font(bold=True)
    _style_header_cell(ws_sum.cell(row=row, column=1))
    _style_header_cell(ws_sum.cell(row=row, column=2))
    row += 1
    risk_counts: dict[str, int] = {}
    for r in scan_results:
        risk_counts[r["t2_risk"]] = risk_counts.get(r["t2_risk"], 0) + 1
    for risk_name in ["LOW", "MEDIUM", "HIGH"]:
        cnt = risk_counts.get(risk_name, 0)
        if cnt > 0:
            ws_sum.cell(row=row, column=1, value=f"T+2 {risk_name}")
            ws_sum.cell(row=row, column=2, value=cnt)
            ws_sum.cell(row=row, column=1).border = _THIN_BORDER
            ws_sum.cell(row=row, column=2).border = _THIN_BORDER
            row += 1

    # Top BUY NOW stocks
    buy_now = [r for r in scan_results if r["recommendation"] == "BUY NOW"]
    if buy_now:
        row += 1
        ws_sum.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws_sum.cell(row=row, column=1, value="Top BUY NOW Stocks").font = Font(
            bold=True, size=12, color="00B050"
        )
        row += 1
        for c, h in enumerate(["Symbol", "LTP", "B/S Ratio", "T+2 Risk", "Score", "Entry Range"], 1):
            cell = ws_sum.cell(row=row, column=c, value=h)
            _style_header_cell(cell)
        row += 1
        for s in buy_now[:10]:
            ws_sum.cell(row=row, column=1, value=s["symbol"])
            ws_sum.cell(row=row, column=2, value=s["live_ltp"])
            ws_sum.cell(row=row, column=3, value=s["buy_sell_ratio"])
            ws_sum.cell(row=row, column=4, value=s["t2_risk"])
            ws_sum.cell(row=row, column=5, value=s["score"])
            ws_sum.cell(row=row, column=6, value=f"{s['entry_low']}-{s['entry_high']}")
            for c in range(1, 7):
                ws_sum.cell(row=row, column=c).fill = _BUY_NOW_FILL
                ws_sum.cell(row=row, column=c).border = _THIN_BORDER
            row += 1

    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 10
    ws_sum.column_dimensions["E"].width = 10
    ws_sum.column_dimensions["F"].width = 16

    wb.save(filepath)
    logger.info(f"Excel saved: {filepath} ({next_row - 2} total rows)")
    return filepath


def _style_header_cell(cell):
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = _THIN_BORDER


def run_live_scan() -> dict:
    """Entry point called by scheduler. Scan + log + append Excel + update cache."""
    global _latest_scan
    now = datetime.now(DSE_TZ)

    # Skip if before 9:55 or after 14:35
    t = now.time()
    from datetime import time as dtime
    if t < dtime(9, 55) or t > dtime(14, 35):
        logger.info(f"Live scanner skipped — outside market hours ({t})")
        return {"skipped": True, "time": str(t)}

    logger.info("Starting live scan...")
    try:
        results = scan_buy_signals()
        if not results:
            logger.warning("No BUY signals to scan")
            return {"scanned": 0}

        filepath = append_to_excel(results)

        # Log decisions to database for backtesting
        _log_decisions(results, now.isoformat())

        # Summarize
        counts: dict[str, int] = {}
        for r in results:
            rec = r["recommendation"]
            counts[rec] = counts.get(rec, 0) + 1

        summary_str = ", ".join(f"{c} {n}" for n, c in sorted(counts.items(), key=lambda x: -x[1]))
        logger.info(f"Live scan complete: {len(results)} stocks — {summary_str}")
        logger.info(f"Excel: {filepath}")

        # Cache for API
        _latest_scan = {
            "timestamp": now.isoformat(),
            "date": now.strftime("%Y-%m-%d"),
            "results": results,
            "summary": counts,
            "total": len(results),
            "excel_path": filepath,
        }

        return _latest_scan

    except Exception as e:
        logger.error(f"Live scan failed: {e}", exc_info=True)
        return {"error": str(e)}


def get_latest_scan() -> dict:
    """Return the latest cached scan results (for API)."""
    return _latest_scan


def get_scan_excel_path(date_str: str | None = None) -> str | None:
    """Return path to the Excel file if it exists."""
    path = _get_excel_path(date_str)
    return path if os.path.exists(path) else None
